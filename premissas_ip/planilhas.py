"""
Extração de premissas a partir das **planilhas de proposição** que alimentam a
engenharia de IP — independente das demais funções do portal (são uploads locais):

- **Extrapolação**: distribuição por classe viária (C0–C5, P5, P6) → Iluminação Viária;
  catálogo de luminárias (potência → custo) como referência.
- **Proposição de IAE**: custo da estrutura de IAE e contagens (pontos atuais/propostos).
- **InvBens (ID / Bens de Interesse)**: lista de Iluminação Especial por bem.

Como o DTO, casa por **palavras-chave nos cabeçalhos** (não por posição fixa), para
funcionar com a planilha de qualquer município no mesmo padrão. Nada é aplicado em
silêncio — tudo aparece para conferência na página.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

from . import dto, schema

_RE_CLASSE = re.compile(r"^(C[0-5]|P[1-6])$", re.IGNORECASE)


def _f(v) -> float | None:
    """Converte célula (número ou texto pt-BR) em float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return dto.num_ptbr(v)


# ── Resultado ─────────────────────────────────────────────────────────────────
@dataclass
class ItemPlan:
    param_id: str
    secao: str
    label_param: str
    origem: str
    valor: float


@dataclass
class ExtracaoPlanilha:
    fonte: str = ""
    valores: dict[str, float] = field(default_factory=dict)
    itens: list[ItemPlan] = field(default_factory=list)
    iluminacao_especial: list[dict] = field(default_factory=list)
    catalogo: list[dict] = field(default_factory=list)
    resultados: list[str] = field(default_factory=list)   # resultados já calculados (read-only)
    avisos: list[str] = field(default_factory=list)

    def _add(self, param_id: str, valor: float, origem: str) -> None:
        s = schema.carregar()
        p = s.parametro(param_id)
        if p is None or param_id in self.valores:
            return
        sec = next((sec.nome for sec in s.secoes if p in sec.parametros), "")
        self.valores[param_id] = valor
        self.itens.append(ItemPlan(param_id, sec, p.label, origem, valor))


# ── Mapa classe → parâmetro (Iluminação Viária do modelo) ─────────────────────
CLASSE_PARAM: dict[str, dict[str, str]] = {
    "Total": {"C0": "p141", "C1": "p142", "C2": "p143", "C3": "p144",
              "P5": "p145", "P6": "p146", "C4": "p147", "C5": "p148"},
    "Vias Principais": {"C0": "p154", "C1": "p155", "C2": "p156", "C3": "p157"},
    "Outras Vias": {"P5": "p160", "P6": "p161", "C4": "p162", "C5": "p163"},
}

# Custo por luminária LED por classe (bloco "Custo Luminária Por Classe" da aba Eco-fin).
CUSTO_CLASSE_PARAM: dict[str, str] = {
    "M1": "p167", "M2": "p168", "M3": "p169", "C0": "p170", "C1": "p171",
    "C2": "p172", "C3": "p173", "P5": "p174", "P6": "p175", "C4": "p176", "C5": "p177",
}

# Rótulos escalares da aba "Inputs Técnicos - Eco-fin" → parâmetro (valor à direita).
LABELS_ECO: tuple[tuple[str, str], ...] = (
    ("custo unitario luminaria", "p285"),    # CPE - custo unitário luminária
    ("custo unitario estrutura", "p286"),    # CPE - custo unitário estrutura
    ("custo medio luminaria vias", "p196"),  # custo médio luminária Vias Principais
    ("custo medio luminaria outras", "p197"),
    ("custo descarte residuos", "p205"),
    ("projeto luminotecnico", "p229"),
    ("custo unitario menor valor", "p225"),  # alterações estruturais (braço)
)


# ── Extrapolação ──────────────────────────────────────────────────────────────
def _achar_distribuicao_classes(wb) -> dict[str, float]:
    """Localiza o bloco 'Classe de iluminação | Quantidade | %' e retorna {classe: fração}."""
    for ws in wb.worksheets:
        linhas = list(ws.iter_rows(values_only=True))
        for ri, row in enumerate(linhas):
            for ci, val in enumerate(row):
                if val and dto._norm(val) == "classe de iluminacao":
                    pct_col = qtd_col = None
                    for cj in range(ci, min(ci + 4, len(row))):
                        h = dto._norm(row[cj]) if row[cj] is not None else ""
                        if h == "%":
                            pct_col = cj
                        elif "quant" in h:
                            qtd_col = cj
                    dist: dict[str, float] = {}
                    quant: dict[str, float] = {}
                    for rj in range(ri + 1, len(linhas)):
                        cell = linhas[rj][ci] if ci < len(linhas[rj]) else None
                        lab = str(cell).strip().upper() if cell is not None else ""
                        if not _RE_CLASSE.match(lab):
                            if dist or quant:
                                break
                            continue
                        if pct_col is not None and pct_col < len(linhas[rj]):
                            v = _f(linhas[rj][pct_col])
                            if v is not None:
                                dist[lab] = v
                        if qtd_col is not None and qtd_col < len(linhas[rj]):
                            q = _f(linhas[rj][qtd_col])
                            if q is not None:
                                quant[lab] = q
                    if not dist and quant:           # só quantidade → calcula fração
                        tot = sum(quant.values())
                        if tot:
                            dist = {k: v / tot for k, v in quant.items()}
                    if len(dist) >= 4:
                        return dist
    return {}


def _catalogo_luminarias(wb) -> list[dict]:
    """Extrai catálogo potência→custo de uma aba 'Banco de dados' (referência)."""
    for ws in wb.worksheets:
        if "banco de dados" not in dto._norm(ws.title):
            continue
        linhas = list(ws.iter_rows(values_only=True))
        for ri, row in enumerate(linhas[:6]):
            norm = [dto._norm(c) if c is not None else "" for c in row]
            pot_col = next((i for i, h in enumerate(norm) if "potencia" in h), None)
            cus_col = next((i for i, h in enumerate(norm) if "custo" in h or "valor" in h), None)
            if pot_col is not None and cus_col is not None:
                cat = []
                for rj in range(ri + 1, len(linhas)):
                    pot = _f(linhas[rj][pot_col]) if pot_col < len(linhas[rj]) else None
                    cus = _f(linhas[rj][cus_col]) if cus_col < len(linhas[rj]) else None
                    if pot and cus:
                        cat.append({"potencia": pot, "custo": cus})
                if cat:
                    return cat
    return []


def _valor_a_direita(row, ci: int) -> float | None:
    """Primeiro valor numérico nas células à direita de `ci` (mesma linha)."""
    for k in (1, 2, 3):
        if ci + k < len(row):
            v = _f(row[ci + k])
            if v is not None:
                return v
    return None


def _extrair_eco_fin(wb, res: ExtracaoPlanilha) -> None:
    """Lê a aba consolidada 'Inputs Técnicos - Eco-fin' da Extrapolação:
    custo de luminária por classe, CPE e custos médios (tudo rotulado)."""
    ws = next((w for w in wb.worksheets if "eco-fin" in dto._norm(w.title)
               or "eco fin" in dto._norm(w.title)), None)
    if ws is None:
        return
    linhas = list(ws.iter_rows(values_only=True))
    for ri, row in enumerate(linhas):
        for ci, val in enumerate(row):
            if val is None:
                continue
            n = dto._norm(val)
            # Bloco "Custo Luminária Por Classe": classe em ci-1, valor em ci.
            if n == "custo luminaria por classe":
                for rj in range(ri + 1, min(ri + 14, len(linhas))):
                    lin = linhas[rj]
                    cl = str(lin[ci - 1]).strip().upper() if ci - 1 >= 0 and lin[ci - 1] else ""
                    if cl in CUSTO_CLASSE_PARAM:
                        v = _f(lin[ci]) if ci < len(lin) else None
                        if v is not None:
                            res._add(CUSTO_CLASSE_PARAM[cl], v, f"Extrapolação · custo luminária {cl} = R$ {v:.2f}")
                    elif cl:               # rótulo não-classe → fim do bloco
                        break
            # Rótulos escalares (valor à direita).
            else:
                for chave, pid in LABELS_ECO:
                    if chave in n:
                        v = _valor_a_direita(row, ci)
                        if v is not None:
                            res._add(pid, v, f"Extrapolação · {val} = {v:.2f}")
                        break


def _extrair_painel_resultados(wb, res: ExtracaoPlanilha) -> None:
    """Lê resultados JÁ calculados pela Extrapolação (aba '10. PainelResultados'):
    eficientização global (→ p119) e expõe potência/consumo/CO2/CPE para conferência."""
    ws = next((w for w in wb.worksheets if "painelresultados" in dto._norm(w.title)
               or "painel resultados" in dto._norm(w.title)), None)
    if ws is None:
        return
    linhas = list(ws.iter_rows(values_only=True))
    rotulos_res = {
        "eficientizacao global": "Eficientização global",
        "reducao da emissao de co2": "Redução de emissão de CO₂",
        "quantidade de correcao de pontos": "Correção de pontos escuros (CPE)",
    }
    for row in linhas:
        for ci, v in enumerate(row):
            if v is None:
                continue
            n = dto._norm(v)
            if "eficientizacao global" in n:
                val = _valor_a_direita(row, ci)
                if val is not None:
                    res._add("p119", val, f"Extrapolação · Eficientização global = {val:.1%}")
            for chave, titulo in rotulos_res.items():
                if chave in n:
                    val = _valor_a_direita(row, ci)
                    if val is not None:
                        res.resultados.append(f"{titulo}: {val:,.2f}".replace(",", "."))
    # Potência total atual → proposta (linha Total da tabela por uso final)
    for row in linhas:
        c0 = dto._norm(row[1]) if len(row) > 1 and row[1] else ""
        if c0 == "total":
            nums = [x for x in row if isinstance(x, (int, float))]
            if len(nums) >= 2:
                res.resultados.append(f"Potência total: {nums[0]:,.0f} → {nums[1]:,.0f} kW"
                                      .replace(",", "."))
                break


def extrair_extrapolacao(arquivo) -> ExtracaoPlanilha:
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    res = ExtracaoPlanilha(fonte="Extrapolação")
    dist = _achar_distribuicao_classes(wb)
    if dist:
        for recorte, mapa in CLASSE_PARAM.items():
            for classe, pid in mapa.items():
                if classe in dist:
                    res._add(pid, dist[classe], f"Extrapolação · {recorte} {classe} = {dist[classe]:.2%}")
    else:
        res.avisos.append("Não encontrei o bloco 'Distribuição de classes' na Extrapolação.")
    _extrair_eco_fin(wb, res)
    _extrair_painel_resultados(wb, res)
    res.catalogo = _catalogo_luminarias(wb)
    if res.catalogo:
        custos = [c["custo"] for c in res.catalogo]
        res.avisos.append(f"Catálogo de luminárias detectado ({len(custos)} itens, "
                          f"custo R$ {min(custos):.0f}–{max(custos):.0f}). Custo por classe requer "
                          "revisão manual (mapeamento classe→potência→custo).")
    wb.close()
    return res


# ── Proposição de IAE ─────────────────────────────────────────────────────────
def _custo_estrutura_iae(wb) -> float | None:
    for ws in wb.worksheets:
        if "estrutura" not in dto._norm(ws.title):
            continue
        linhas = list(ws.iter_rows(values_only=True))
        for ri, row in enumerate(linhas[:4]):
            for ci, val in enumerate(row):
                if val and "custo un" in dto._norm(val):
                    for rj in range(ri + 1, len(linhas)):
                        v = _f(linhas[rj][ci]) if ci < len(linhas[rj]) else None
                        if v:
                            return v
    return None


def _iae_totais(wb) -> dict[str, float]:
    """Encontra a linha TOTAL do pivot atual×proposto da Proposição IAE."""
    for ws in wb.worksheets:
        linhas = list(ws.iter_rows(values_only=True))
        for ri, row in enumerate(linhas):
            norm = [dto._norm(c) if c is not None else "" for c in row]
            atual_col = next((i for i, h in enumerate(norm)
                              if "quantidade de pontos" in h and "propos" not in h), None)
            total_col = next((i for i, h in enumerate(norm) if "quantidade total de pontos" in h), None)
            if atual_col is not None and total_col is not None:
                for rj in range(ri + 1, len(linhas)):
                    c0 = linhas[rj][0] if linhas[rj] else None
                    if c0 and dto._norm(c0) in ("total", "total geral"):
                        return {
                            "atuais": _f(linhas[rj][atual_col]) or 0,
                            "total": _f(linhas[rj][total_col]) or 0,
                        }
    return {}


def extrair_proposicao_iae(arquivo) -> ExtracaoPlanilha:
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    res = ExtracaoPlanilha(fonte="Proposição de IAE")
    custo = _custo_estrutura_iae(wb)
    if custo:
        res._add("p248", custo, f"Proposição IAE · Estruturas = R$ {custo:.2f}")
    tot = _iae_totais(wb)
    if tot:
        res._add("p251", tot["total"], f"Proposição IAE · pontos propostos (total) = {tot['total']:.0f}")
        res.avisos.append(f"IAE detectado: {tot['atuais']:.0f} pontos atuais → {tot['total']:.0f} propostos. "
                          "Confirme o mapeamento (atual × Fase de Modernização × Demanda Reprimida).")
    if not res.valores:
        res.avisos.append("Não reconheci custo de estrutura nem contagens na Proposição de IAE.")
    wb.close()
    return res


# ── InvBens (ID / Bens de Interesse) → lista de Iluminação Especial ───────────
def _capex_por_bem(wb) -> dict[str, float]:
    """Soma o CAPEX por bem na aba 'Quadro Investimentos' (coluna Total agrupada por Bem)."""
    ws = next((w for w in wb.worksheets if "quadro investimento" in dto._norm(w.title)), None)
    if ws is None:
        return {}
    linhas = list(ws.iter_rows(values_only=True))
    bem_col = total_col = hdr = None
    for ri, row in enumerate(linhas[:20]):
        norm = [dto._norm(c) if c is not None else "" for c in row]
        bc = next((i for i, h in enumerate(norm) if h == "bem"), None)
        tc = next((i for i, h in enumerate(norm) if h == "total"), None)
        if bc is not None and tc is not None:
            bem_col, total_col, hdr = bc, tc, ri
            break
    if bem_col is None:
        return {}
    capex: dict[str, float] = {}
    for rj in range(hdr + 1, len(linhas)):
        bem = linhas[rj][bem_col] if bem_col < len(linhas[rj]) else None
        if not bem or not str(bem).strip():
            continue
        v = _f(linhas[rj][total_col]) if total_col < len(linhas[rj]) else None
        if v:
            capex[dto._norm(bem)] = capex.get(dto._norm(bem), 0.0) + v
    return capex


def extrair_invbens(arquivo) -> ExtracaoPlanilha:
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    res = ExtracaoPlanilha(fonte="InvBens (ID)")
    ws = next((w for w in wb.worksheets if "planilha de engenharia" in dto._norm(w.title)), None)
    if ws is None:
        res.avisos.append("Aba 'Planilha de Engenharia' não encontrada no InvBens.")
        wb.close()
        return res
    linhas = list(ws.iter_rows(values_only=True))
    bem_col = lum_col = hdr = None
    for ri, row in enumerate(linhas[:12]):
        norm = [dto._norm(c) if c is not None else "" for c in row]
        bc = next((i for i, h in enumerate(norm) if h == "bem"), None)
        lc = next((i for i, h in enumerate(norm) if "luminarias existentes" in h or "luminarias" in h), None)
        if bc is not None:
            bem_col, lum_col, hdr = bc, lc, ri
            break
    if bem_col is None:
        res.avisos.append("Cabeçalho 'Bem' não localizado no InvBens.")
        wb.close()
        return res
    acc: dict[str, float] = {}
    for rj in range(hdr + 1, len(linhas)):
        bem = linhas[rj][bem_col] if bem_col < len(linhas[rj]) else None
        if not bem or not str(bem).strip():
            continue
        nome = str(bem).strip()
        lum = _f(linhas[rj][lum_col]) if (lum_col is not None and lum_col < len(linhas[rj])) else None
        acc[nome] = max(acc.get(nome, 0.0), lum or 0.0)
    capex = _capex_por_bem(wb)
    res.iluminacao_especial = [
        {"local": nome, "pontos_atuais": int(p), "pontos_futuros": 0,
         "capex": round(capex.get(dto._norm(nome), 0.0), 2)}
        for nome, p in acc.items()
    ]
    n_capex = sum(1 for it in res.iluminacao_especial if it["capex"])
    if n_capex:
        res.avisos.append(f"CAPEX por bem somado do 'Quadro Investimentos' para {n_capex} item(ns).")
    if not res.iluminacao_especial:
        res.avisos.append("Nenhum bem encontrado na aba de engenharia do InvBens.")
    wb.close()
    return res


__all__ = [
    "ExtracaoPlanilha", "ItemPlan", "CLASSE_PARAM",
    "extrair_extrapolacao", "extrair_proposicao_iae", "extrair_invbens",
]
