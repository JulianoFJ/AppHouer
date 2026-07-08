"""
Gerador dos **blocos de dados do relatório de engenharia**.

Produz (1) um .xlsx com as tabelas que entram no relatório — Resumo CAPEX (§6),
Resumo OPEX (§7) e Quantitativos — e (2) um texto .md com orientação curta de como
cada número é obtido. Cada linha traz uma coluna **Memória de cálculo** com a fórmula
em linguagem natural e os parâmetros de origem, para rastreabilidade/auditoria.

Os números são calculados a partir das `Respostas`; a parametrização "viva" (fórmulas
Excel recalculáveis) está na planilha de inputs (`planilha_inputs.py`). Aqui o foco é o
bloco pronto para o relatório, com a conta documentada.
"""

from __future__ import annotations

from openpyxl import Workbook

from cadastro_ip.saidas import _helpers as H

from .. import modelo, schema
from ..coleta import Respostas

FMT_MOEDA = 'R$ #,##0.00'


def _v(resp: Respostas, pid: str, padrao: float = 0.0) -> float:
    """Valor numérico de um parâmetro (coletado ou default do catálogo)."""
    raw = resp.valores.get(pid, None)
    if raw is None or raw == "":
        p = schema.carregar().parametro(pid)
        raw = p.default if p else None
    if raw is None or raw == "":
        return padrao
    try:
        return float(str(raw).replace(",", "."))
    except (TypeError, ValueError):
        return padrao


def _capex(resp: Respostas) -> list[tuple[str, float, str]]:
    """Linhas de CAPEX: (item, valor, memória de cálculo)."""
    parque = _v(resp, modelo.PARQUE_TOTAL)
    preco_lum = _v(resp, "p195")
    tele_capex = _v(resp, modelo.TELEGESTAO_CAPEX_PONTO)
    exp_anual = _v(resp, modelo.EXP_ANUAL_PONTOS)
    custo_lum_exp = _v(resp, modelo.CUSTO_MEDIO_LUM_EXP)
    capex_ie = sum(float(it.get("capex", 0) or 0) for it in resp.iluminacao_especial)

    return [
        ("Modernização de Luminárias", parque * preco_lum,
         "Parque total (p37) × Preço médio luminária LED (p195)"),
        ("Sistema de Telegestão", parque * tele_capex,
         "Parque total (p37) × CAPEX telegestão por ponto (p241)"),
        ("Expansão (1º ano)", exp_anual * custo_lum_exp,
         "Expansão anual (p349) × Custo médio luminária expansão (p296)"),
        ("Iluminação Especial", capex_ie,
         "Soma do CAPEX dos itens listados na seção Iluminação Especial"),
        ("CAPEX de veículos", _v(resp, modelo.VEIC_QTD_CESTO) * _v(resp, modelo.VEIC_VAL_CESTO)
         + _v(resp, modelo.VEIC_QTD_MOTO) * _v(resp, modelo.VEIC_VAL_MOTO),
         "Σ quantidade × valor (cesto p686×p689 + moto p687×p690)"),
    ]


def _opex(resp: Respostas) -> list[tuple[str, float, str]]:
    """Linhas de OPEX mensal: (item, valor, memória de cálculo)."""
    enc = _v(resp, modelo.ENCARGOS)
    benef = _v(resp, modelo.BENEFICIOS)
    sumprod = sum(_v(resp, c) * _v(resp, s) for c, s in modelo.COMPOSICAO_SALARIO)
    soma_comp = sum(_v(resp, c) for c, _ in modelo.COMPOSICAO_SALARIO)
    custo_equipe = sumprod * (1 + enc) + soma_comp * benef

    frota_setup = _v(resp, modelo.FROTA_MARCO["setup"])
    opex_veic = frota_setup * (_v(resp, modelo.VEIC_LOCACAO_CESTO)
                               + _v(resp, modelo.VEIC_COMBUSTIVEL_CESTO)
                               + _v(resp, modelo.VEIC_DESPESAS_CESTO))
    parque = _v(resp, modelo.PARQUE_TOTAL)
    tele_opex = parque * _v(resp, modelo.TELEGESTAO_OPEX_PONTO)
    poda = _v(resp, "p751")

    return [
        ("Equipe operacional (1 equipe)", custo_equipe,
         "Σ(composição×salário)×(1+encargos p641) + Σ(composição)×benefícios (p639)"),
        ("Veículos (frota setup)", opex_veic,
         "Frota do marco setup (p694) × (locação p706 + combustível p712 + despesas p715)"),
        ("Telegestão (licenciamento)", tele_opex,
         "Parque total (p37) × OPEX telegestão por ponto (p242)"),
        ("Poda de árvore (mensal)", poda,
         "Valor total/mês de poda (p751)"),
    ]


def _escrever_bloco(ws, titulo: str, linhas: list[tuple[str, float, str]], linha0: int) -> int:
    ws.cell(row=linha0, column=1, value=titulo)
    H.aplicar_estilo_subheader(ws, linha0, 3)
    r = linha0 + 1
    for col, txt in enumerate(["Item", "Valor (R$)", "Memória de cálculo"], start=1):
        ws.cell(row=r, column=col, value=txt)
    H.aplicar_estilo_header(ws, r, 3)
    r += 1
    total = 0.0
    for item, valor, memoria in linhas:
        ws.cell(row=r, column=1, value=item)
        c = ws.cell(row=r, column=2, value=round(valor, 2))
        c.number_format = FMT_MOEDA
        ws.cell(row=r, column=3, value=memoria)
        total += valor
        r += 1
    ws.cell(row=r, column=1, value="TOTAL")
    c = ws.cell(row=r, column=2, value=round(total, 2))
    c.number_format = FMT_MOEDA
    H.aplicar_estilo_total(ws, r, 3)
    return r + 2


def gerar(resp: Respostas) -> tuple[bytes, str]:
    """Gera (bytes_xlsx, texto_md) dos blocos do relatório."""
    capex = _capex(resp)
    opex = _opex(resp)

    wb = Workbook()
    ws = wb.active
    ws.title = "Blocos Relatório"
    ws.cell(row=1, column=1, value=f"Blocos do Relatório de Engenharia — {resp.municipio or 'Município'}/{resp.uf}")
    H.aplicar_estilo_header(ws, 1, 3)
    prox = _escrever_bloco(ws, "Resumo CAPEX (§6)", capex, 3)
    prox = _escrever_bloco(ws, "Resumo OPEX mensal (§7)", opex, prox)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70
    xlsx = H.workbook_para_bytes(wb)

    total_capex = sum(v for _, v, _ in capex)
    total_opex = sum(v for _, v, _ in opex)
    linhas_md = [
        f"# Blocos do Relatório de Engenharia — {resp.municipio or 'Município'}/{resp.uf}",
        "",
        "> Estes blocos alimentam as seções §6 (Investimentos/CAPEX) e §7 (Custos/OPEX) "
        "do relatório. Os valores são derivados das premissas coletadas; a coluna "
        "*Memória de cálculo* na planilha documenta cada conta, e a planilha de inputs "
        "traz as mesmas relações como fórmulas vivas (recalculáveis no Excel).",
        "",
        "## Resumo CAPEX (§6)",
        "",
        "| Item | Valor (R$) | Memória de cálculo |",
        "|---|---:|---|",
    ]
    for item, valor, memoria in capex:
        linhas_md.append(f"| {item} | {valor:,.2f} | {memoria} |")
    linhas_md.append(f"| **TOTAL** | **{total_capex:,.2f}** | |")
    linhas_md += ["", "## Resumo OPEX mensal (§7)", "",
                  "| Item | Valor (R$) | Memória de cálculo |", "|---|---:|---|"]
    for item, valor, memoria in opex:
        linhas_md.append(f"| {item} | {valor:,.2f} | {memoria} |")
    linhas_md.append(f"| **TOTAL** | **{total_opex:,.2f}** | |")
    md = "\n".join(linhas_md)

    return xlsx, md


__all__ = ["gerar"]
