"""
Gerador da **Planilha de Inputs IP** parametrizada.

Produz um .xlsx em memória com:
- aba `Inputs_IP`: as 32 seções no layout do modelo (cada parâmetro na sua linha
  original → as referências do modelo econômico-financeiro continuam válidas),
  com unidade, valor coletado e Fonte.
- aba `Dimensionamento`: bases de CAPEX/OPEX e custo de equipe, como fórmulas que
  referenciam as células de `Inputs_IP`.
- aba `Distribuição Temporal`: linha do tempo ano a ano (1..prazo de concessão) de
  expansão e reinvestimento por vida útil — tudo em fórmulas vivas e auditáveis.

Reusa os helpers de estilo/escrita de `cadastro_ip.saidas._helpers`.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cadastro_ip.saidas import _helpers as H

from .. import coleta, modelo, schema
from ..coleta import Respostas

# Colunas (1-based) que reproduzem a geometria do modelo.
C_MARCA, C_LABEL, C_UNID, C_VALOR, C_FONTE = 1, 2, 3, 6, 12

FMT = {
    "percentual": "0.0000%",
    "moeda": 'R$ #,##0.00',
    "numero": "#,##0.00",
}


def _valor(resp: Respostas, p: schema.Parametro):
    """Valor a gravar para um parâmetro: coletado se houver, senão default do catálogo."""
    v = resp.valores.get(p.id, None)
    if v is None or v == "":
        v = p.default
    if v is None or v == "":
        return None
    if p.tipo in ("numero", "moeda", "percentual"):
        try:
            return float(str(v).replace(",", "."))
        except (TypeError, ValueError):
            return None
    return v


def _aba_inputs(wb: Workbook, resp: Respostas) -> None:
    ws = wb.active
    ws.title = modelo.ABA_INPUTS
    s = schema.carregar()

    # Cabeçalho de colunas na linha 1 (vazia no modelo → não desloca referências).
    ws.cell(row=1, column=C_LABEL, value="Parâmetro")
    ws.cell(row=1, column=C_UNID, value="Unidade")
    ws.cell(row=1, column=C_VALOR, value="Valor")
    ws.cell(row=1, column=C_FONTE, value="Fonte")
    H.aplicar_estilo_header(ws, 1, C_FONTE)

    for sec in s.secoes:
        r = sec.linha_modelo
        # Seções fora do escopo da engenharia (Financeiro/Socioambiental): mantém a
        # estrutura no layout, mas com valores vazios (preenchidas pela equipe eco-fin).
        escopo = coleta.secao_no_escopo(sec)
        ws.cell(row=r, column=C_MARCA, value="x")
        ws.cell(row=r, column=C_LABEL, value=sec.nome)
        H.aplicar_estilo_subheader(ws, r, C_LABEL, colunas_a_estilizar=[C_MARCA, C_LABEL])

        for p in sec.parametros:
            rr = p.linha_modelo
            ws.cell(row=rr, column=C_LABEL, value=p.label)
            if p.eh_grupo:
                ws.cell(row=rr, column=C_LABEL).font = H.total_font()
                continue
            if p.unidade:
                ws.cell(row=rr, column=C_UNID, value=p.unidade)
            val = _valor(resp, p) if escopo else None
            if val is not None:
                cel = ws.cell(row=rr, column=C_VALOR, value=val)
                if p.tipo in FMT and isinstance(val, (int, float)):
                    cel.number_format = FMT[p.tipo]
            if p.fonte:
                ws.cell(row=rr, column=C_FONTE, value=p.fonte).font = H.body_font()

    ws.column_dimensions[get_column_letter(C_LABEL)].width = 52
    ws.column_dimensions[get_column_letter(C_UNID)].width = 14
    ws.column_dimensions[get_column_letter(C_VALOR)].width = 18
    ws.column_dimensions[get_column_letter(C_FONTE)].width = 46
    ws.freeze_panes = "B2"


def _aba_dimensionamento(wb: Workbook) -> dict[str, str]:
    """Bases de custo como fórmulas. Retorna mapa nome->referência de célula (p/ outras abas)."""
    ws = wb.create_sheet("Dimensionamento")
    ws.cell(row=1, column=1, value="Dimensionamento — fórmulas referenciam a aba Inputs_IP")
    H.aplicar_estilo_header(ws, 1, 4)

    linhas = [
        ("Custo mensal de uma equipe operacional (R$)", modelo.f_custo_equipe_operacional()),
        ("CAPEX aquisição de veículos (R$)", modelo.f_capex_veiculos()),
        ("CAPEX Telegestão (R$)", modelo.f_capex_telegestao()),
    ]
    if modelo.existe("p195") and modelo.existe(modelo.PARQUE_TOTAL):
        linhas.append(("CAPEX Modernização Luminárias (R$)",
                       f"={modelo.cel(modelo.PARQUE_TOTAL)}*{modelo.cel('p195')}"))

    ref: dict[str, str] = {}
    r = 3
    for rotulo, formula in linhas:
        ws.cell(row=r, column=2, value=rotulo)
        c = ws.cell(row=r, column=3, value=formula)
        c.number_format = FMT["moeda"]
        ref[rotulo] = f"Dimensionamento!$C${r}"
        r += 1

    # Custo operacional por marco (referencia a célula de custo de equipe).
    cel_equipe = ref["Custo mensal de uma equipe operacional (R$)"]
    r += 1
    ws.cell(row=r, column=2, value="Custo operacional mensal por marco (R$)")
    ws.cell(row=r, column=2).font = H.total_font()
    r += 1
    for marco in modelo.EQ_MARCO:
        ws.cell(row=r, column=2, value=f"Marco {marco}")
        c = ws.cell(row=r, column=3, value=modelo.f_custo_operacional_marco(marco, cel_equipe))
        c.number_format = FMT["moeda"]
        r += 1

    # OPEX de veículos por marco.
    r += 1
    ws.cell(row=r, column=2, value="OPEX mensal de veículos por marco (R$)")
    ws.cell(row=r, column=2).font = H.total_font()
    r += 1
    for marco in modelo.FROTA_MARCO:
        ws.cell(row=r, column=2, value=f"Marco {marco}")
        c = ws.cell(row=r, column=3, value=modelo.f_opex_veiculos_marco(marco))
        c.number_format = FMT["moeda"]
        r += 1

    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 20
    return ref


def _aba_distribuicao_temporal(wb: Workbook, ref: dict[str, str]) -> None:
    ws = wb.create_sheet("Distribuição Temporal")
    n = modelo.anos_concessao_default()

    ws.cell(row=1, column=1,
            value="Distribuição temporal de custos (R$) — expansão e reinvestimento por vida útil. "
                  "Fórmulas vivas referenciando Inputs_IP / Dimensionamento.")
    # Linha de anos.
    ws.cell(row=2, column=1, value="Item \\ Ano")
    for a in range(1, n + 1):
        ws.cell(row=2, column=1 + a, value=a)
    H.aplicar_estilo_header(ws, 2, 1 + n)

    # Pontos expandidos por ano.
    r_pontos = 3
    ws.cell(row=r_pontos, column=1, value="Pontos expandidos no ano")
    for a in range(1, n + 1):
        ws.cell(row=r_pontos, column=1 + a, value=modelo.f_expansao_pontos_ano(a))

    # CAPEX de expansão por ano (referencia a célula de pontos do mesmo ano).
    r_capex_exp = 4
    ws.cell(row=r_capex_exp, column=1, value="CAPEX Expansão")
    for a in range(1, n + 1):
        cel_pontos = f"{get_column_letter(1 + a)}{r_pontos}"
        c = ws.cell(row=r_capex_exp, column=1 + a, value=modelo.f_capex_expansao_ano(cel_pontos))
        c.number_format = FMT["moeda"]

    # Reinvestimento de luminárias (vida útil) — base = CAPEX modernização luminárias.
    base_lum = ref.get("CAPEX Modernização Luminárias (R$)")
    if base_lum:
        r = 5
        ws.cell(row=r, column=1, value="Reinvestimento Luminárias")
        for a in range(1, n + 1):
            c = ws.cell(row=r, column=1 + a,
                        value=modelo.f_reinvestimento_ano(base_lum, modelo.cel(modelo.VIDA_UTIL_LUMINARIA), a))
            c.number_format = FMT["moeda"]

    # Reinvestimento de telegestão.
    base_tele = ref.get("CAPEX Telegestão (R$)")
    if base_tele:
        r = 6
        ws.cell(row=r, column=1, value="Reinvestimento Telegestão")
        for a in range(1, n + 1):
            c = ws.cell(row=r, column=1 + a,
                        value=modelo.f_reinvestimento_ano(base_tele, modelo.cel(modelo.VIDA_UTIL_TELEGESTAO), a))
            c.number_format = FMT["moeda"]

    ws.column_dimensions["A"].width = 28
    ws.freeze_panes = "B3"


def gerar(resp: Respostas) -> bytes:
    """Monta a planilha de inputs parametrizada e devolve os bytes do .xlsx."""
    wb = Workbook()
    _aba_inputs(wb, resp)
    ref = _aba_dimensionamento(wb)
    _aba_distribuicao_temporal(wb, ref)
    return H.workbook_para_bytes(wb)


__all__ = ["gerar"]
