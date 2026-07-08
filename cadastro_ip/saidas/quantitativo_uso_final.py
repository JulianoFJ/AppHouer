"""
Gerador do arquivo `Quantitativo por Uso Final.xlsx` (seção 12.3).

Estrutura (replicando o modelo ARMBH):
  - Aba `TOTAL` — consolidação por **tecnologia** (somando IV + IAE + ID),
    com Carga e Consumo. (No modelo agrega por tecnologia apenas.)
  - Aba `Resumo IV IAE e ID` — três tabelas lado a lado (uma por uso final),
    cada uma com granularidade Tec × Pot.
  - Aba `Inventário Total` — lista por uso final, mantendo Tec × Pot.
  - Abas `Base IAE`, `Base ID`, `Base IV` — detalhamento por uso final.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ._helpers import (
    aplicar_estilo_header,
    aplicar_estilo_subheader,
    aplicar_estilo_total,
    autoajustar_largura,
    escrever_dataframe,
    gerar_combinacoes_tec_pot,
    workbook_para_bytes,
)


def gerar(resultado) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # Construir o universo consolidado (IV + IAE + ID após tratamentos)
    inventario = _construir_inventario(resultado)

    _aba_inventario_total(wb, inventario, resultado)
    _aba_base_iv(wb, resultado, inventario)
    _aba_base_iae(wb, resultado, inventario)
    _aba_base_id(wb, resultado, inventario)
    _aba_resumo_iv_iae_id(wb, inventario, resultado)
    _aba_total(wb, inventario, resultado)

    # Reordena: TOTAL deve aparecer primeiro (padrão do modelo)
    nova_ordem = ["TOTAL", "Resumo IV IAE e ID", "Inventário Total", "Base IAE", "Base ID", "Base IV"]
    wb._sheets.sort(key=lambda s: nova_ordem.index(s.title) if s.title in nova_ordem else 99)

    return workbook_para_bytes(wb)


# ── Construção do inventário consolidado ──────────────────────────────────────
def _construir_inventario(r) -> pd.DataFrame:
    """
    Inventário consolidado pós-tratamento: cada ponto do Cadastro Corrigido
    (cadastro original + IAE novos + ID novos) com sua Tecnologia/Potência
    finais (após regra Considerada) e o uso final atribuído.
    """
    cad = r.cadastro_normalizado.copy()
    cad["uso_final"] = "IV"

    # Reatribui IAE / ID a partir dos caminhos
    rot = r.roteamento
    if not rot.df_roteado.empty and "caminho" in rot.df_roteado.columns:
        cad_com_caminho = rot.df_roteado[["id_ponto", "caminho"]].rename(columns={"caminho": "_caminho"})
        cad = cad.merge(cad_com_caminho, on="id_ponto", how="left")
        cad.loc[cad["_caminho"] == "IAE_EXISTENTE", "uso_final"] = "IAE"
        cad.loc[cad["_caminho"] == "ID_EXISTENTE", "uso_final"] = "ID"
        cad = cad.drop(columns=["_caminho"])

    # Adiciona pontos novos de IAE e ID
    iae_novos = rot.pontos_iae_novos.copy()
    if not iae_novos.empty:
        iae_novos["uso_final"] = "IAE"
    id_novos = rot.pontos_id_novos.copy()
    if not id_novos.empty:
        id_novos["uso_final"] = "ID"

    # União: garantir colunas comuns
    cols_essenciais = ["id_ponto", "logradouro", "local", "latitude", "longitude",
                       "codigo_tecnologia", "potencia", "Quantidade", "quantidade",
                       "reator_w", "uso_final"]
    def _normalizar(d):
        for c in cols_essenciais:
            if c not in d.columns:
                d[c] = None
        return d[cols_essenciais]

    inventario = pd.concat([_normalizar(cad), _normalizar(iae_novos), _normalizar(id_novos)], ignore_index=True)

    # Unifica qty (Quantidade vs quantidade). Arredonda antes do astype(int) para evitar
    # truncamento de valores como 0.999999... (float storage de quantidade inteira).
    inventario["qtd"] = pd.to_numeric(
        inventario["quantidade"].fillna(inventario["Quantidade"]).fillna(1), errors="coerce"
    ).fillna(1).round().astype(int)
    inventario = inventario.drop(columns=["Quantidade", "quantidade"])
    inventario = inventario.rename(columns={"qtd": "quantidade"})

    return inventario


# ── Aba Inventário Total ──────────────────────────────────────────────────────
def _aba_inventario_total(wb: Workbook, inventario: pd.DataFrame, r) -> None:
    ws = wb.create_sheet("Inventário Total")
    df = inventario.copy()
    df["potencia_total"] = (
        pd.to_numeric(df["potencia"], errors="coerce").fillna(0) +
        pd.to_numeric(df["reator_w"], errors="coerce").fillna(0)
    ) * df["quantidade"]
    df = df.rename(columns={
        "id_ponto": "ID",
        "uso_final": "Uso Final",
        "local": "Tipo de Local",
        "logradouro": "Logradouro",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "codigo_tecnologia": "Tecnologia",
        "potencia": "Potência",
        "reator_w": "Reator",
        "quantidade": "Quantidade",
        "potencia_total": "Potência Total",
    })
    ordem = ["Uso Final", "ID", "Tipo de Local", "Logradouro", "Latitude", "Longitude",
             "Tecnologia", "Potência", "Reator", "Quantidade", "Potência Total"]
    df = df[[c for c in ordem if c in df.columns]]
    escrever_dataframe(ws, df, linha_inicial=1)
    autoajustar_largura(ws, len(df.columns))
    ws.freeze_panes = "A2"


# ── Abas Base IV / IAE / ID ───────────────────────────────────────────────────
def _aba_base_iv(wb: Workbook, r, inventario: pd.DataFrame) -> None:
    _aba_base_por_uso(wb, "Base IV", "IV", inventario, r)


def _aba_base_iae(wb: Workbook, r, inventario: pd.DataFrame) -> None:
    _aba_base_por_uso(wb, "Base IAE", "IAE", inventario, r)


def _aba_base_id(wb: Workbook, r, inventario: pd.DataFrame) -> None:
    _aba_base_por_uso(wb, "Base ID", "ID", inventario, r)


def _aba_base_por_uso(wb: Workbook, nome_aba: str, uso: str, inventario: pd.DataFrame, r) -> None:
    """
    Tabela Tec × Pot do uso especificado, com valores pré-calculados em Python.
    Apenas a coluna "Potência Total" mantém fórmula viva (Pot + Reator) × Qtd
    para o Excel recalcular se o usuário editar manualmente.
    Bloco de Tempo de Operação fica isolado (cols H/I) sem sobrepor estilos.
    """
    from ..reator import perda_reator

    ws = wb.create_sheet(nome_aba)
    subset = inventario[inventario["uso_final"] == uso]
    combos = gerar_combinacoes_tec_pot(subset, "codigo_tecnologia", "potencia")

    # Cabeçalho da tabela principal (cols A-E)
    headers = ["Tecnologia", "Potência (W)", "Reator (W)", "Quantidade", "Potência Total (W)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    aplicar_estilo_header(ws, 1, 5)

    # Bloco lateral de Tempo de Operação — completamente desacoplado da tabela
    ws.cell(row=1, column=8, value="Tempo de Operação")
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
    aplicar_estilo_header(ws, 1, 9)
    ws.cell(row=2, column=8, value="Horas")
    ws.cell(row=2, column=9, value="Minutos")
    aplicar_estilo_subheader(ws, 2, 9, colunas_a_estilizar=[8, 9])
    if r.tempo_operacao is not None:
        ws.cell(row=3, column=8, value=int(r.tempo_operacao.horas))
        ws.cell(row=3, column=9, value=int(r.tempo_operacao.minutos))

    if combos.empty:
        ws.cell(row=2, column=1, value=f"Sem pontos no uso final {uso}.")
        autoajustar_largura(ws, 9, largura_minima=12)
        return

    primeira = 2
    ultima = primeira + len(combos) - 1

    for i, row in combos.iterrows():
        r_ = primeira + i
        tec = str(row["codigo_tecnologia"]) if pd.notna(row["codigo_tecnologia"]) else ""
        pot = float(row["potencia"]) if pd.notna(row["potencia"]) else 0.0
        qtd = int(round(float(row["quantidade"]))) if pd.notna(row["quantidade"]) else 0
        is_led = tec.upper() == "LED"
        reator = float(perda_reator(pot, is_led=is_led))

        ws.cell(row=r_, column=1, value=tec)
        ws.cell(row=r_, column=2, value=pot)
        ws.cell(row=r_, column=3, value=reator)
        ws.cell(row=r_, column=4, value=qtd)
        ws.cell(row=r_, column=5, value=f"=(B{r_}+C{r_})*D{r_}")

    # Linha TOTAL (uma linha em branco depois dos dados, padrão modelo)
    linha_total = ultima + 2
    ws.cell(row=linha_total, column=1, value="TOTAL")
    for col_idx in [4, 5]:
        col_letra = get_column_letter(col_idx)
        ws.cell(row=linha_total, column=col_idx,
                value=f"=SUM({col_letra}{primeira}:{col_letra}{ultima})")
    aplicar_estilo_total(ws, linha_total, 5)

    autoajustar_largura(ws, 9, largura_minima=12)
    ws.freeze_panes = "A2"


# ── Aba Resumo IV IAE e ID (3 tabelas lado a lado) ───────────────────────────
def _aba_resumo_iv_iae_id(wb: Workbook, inventario: pd.DataFrame, r) -> None:
    ws = wb.create_sheet("Resumo IV IAE e ID")
    blocos = [("IV", 1), ("IAE", 7), ("ID", 13)]  # col inicial para cada bloco
    for uso, col_inicio in blocos:
        ws.cell(row=1, column=col_inicio, value="Uso Final")
        ws.cell(row=1, column=col_inicio + 1, value=uso)
        ws.merge_cells(start_row=1, start_column=col_inicio + 1, end_row=1, end_column=col_inicio + 4)

        headers = ["Tecnologia", "Potência (W)", "Reator (W)", "Quantidade", "Potência Total (W)"]
        for i, h in enumerate(headers):
            ws.cell(row=3, column=col_inicio + i, value=h)

        subset = inventario[inventario["uso_final"] == uso]
        combos = gerar_combinacoes_tec_pot(subset, "codigo_tecnologia", "potencia")

        if combos.empty:
            ws.cell(row=4, column=col_inicio, value=f"Sem pontos no uso {uso}.")
            continue

        primeira = 4
        for i, row in combos.iterrows():
            r_ = primeira + i
            ws.cell(row=r_, column=col_inicio + 0, value=row["codigo_tecnologia"])
            ws.cell(row=r_, column=col_inicio + 1, value=float(row["potencia"]) if pd.notna(row["potencia"]) else 0)
            # Para o resumo usamos valores chumbados (não há aba auxiliar p/ SUMIFS aqui)
            from ..reator import perda_reator
            is_led = str(row["codigo_tecnologia"]).upper() == "LED"
            reator_w = perda_reator(row["potencia"], is_led=is_led)
            ws.cell(row=r_, column=col_inicio + 2, value=float(reator_w))
            ws.cell(row=r_, column=col_inicio + 3, value=int(row["quantidade"]))
            # Potência Total — fórmula viva referenciando colunas do mesmo bloco
            col_pot = get_column_letter(col_inicio + 1)
            col_reat = get_column_letter(col_inicio + 2)
            col_qtd = get_column_letter(col_inicio + 3)
            ws.cell(row=r_, column=col_inicio + 4, value=f"=({col_pot}{r_}+{col_reat}{r_})*{col_qtd}{r_}")

        # Linha TOTAL
        ultima = primeira + len(combos) - 1
        linha_total = ultima + 1
        ws.cell(row=linha_total, column=col_inicio, value="Total Geral")
        col_qtd_letra = get_column_letter(col_inicio + 3)
        col_pt_letra = get_column_letter(col_inicio + 4)
        ws.cell(row=linha_total, column=col_inicio + 3, value=f"=SUM({col_qtd_letra}{primeira}:{col_qtd_letra}{ultima})")
        ws.cell(row=linha_total, column=col_inicio + 4, value=f"=SUM({col_pt_letra}{primeira}:{col_pt_letra}{ultima})")
        aplicar_estilo_total(ws, linha_total, col_inicio + 4)

    aplicar_estilo_header(ws, 1, 17)
    aplicar_estilo_subheader(ws, 3, 17)
    autoajustar_largura(ws, 17, largura_minima=11)
    ws.freeze_panes = "A4"


# ── Aba TOTAL (consolidado por tecnologia + Carga + Consumo) ─────────────────
def _aba_total(wb: Workbook, inventario: pd.DataFrame, r) -> None:
    """
    Aba TOTAL replica o modelo ARMBH: agrega por TECNOLOGIA apenas (não Tec × Pot).
    Estrutura:
      Bloco superior: Tempo de Consumo Diário (Horas | Minutos)
      Tabela:
        Tecnologia | Sum of Potência Total | Sum of Quantidade |
        Representatividade | Carga [kW] | Consumo [kWh]
      Linha Total Geral ao final.
    """
    ws = wb.create_sheet("TOTAL")

    # Tempo de Consumo Diário
    ws["B1"] = "Tempo de Consumo Diário"
    ws.merge_cells("B1:C1")
    ws["B2"] = "Horas"
    ws["C2"] = "Minutos"
    if r.tempo_operacao is not None:
        ws["B3"] = int(r.tempo_operacao.horas)
        ws["C3"] = int(r.tempo_operacao.minutos)

    aplicar_estilo_header(ws, 1, 7)
    aplicar_estilo_subheader(ws, 2, 7)

    # Agrega por tecnologia
    df = inventario.copy()
    df["potencia"] = pd.to_numeric(df["potencia"], errors="coerce").fillna(0)
    df["reator_w"] = pd.to_numeric(df["reator_w"], errors="coerce").fillna(0)
    df["potencia_total"] = (df["potencia"] + df["reator_w"]) * df["quantidade"]
    agg = df.groupby("codigo_tecnologia", dropna=False).agg(
        potencia_total=("potencia_total", "sum"),
        quantidade=("quantidade", "sum"),
    ).reset_index()
    # Ordena por tecnologia (LED primeiro, depois VS, VM, ...)
    from ._helpers import _ordem_tecnologia
    agg["_ord"] = agg["codigo_tecnologia"].map(_ordem_tecnologia)
    agg = agg.sort_values(["_ord", "codigo_tecnologia"]).drop(columns=["_ord"]).reset_index(drop=True)

    # Cabeçalho da tabela na linha 5
    headers = ["", "Tecnologia", "Sum of Potência Total", "Sum of Quantidade",
               "Representatividade", "Carga [kW]", "Consumo [kWh/mês]"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    aplicar_estilo_header(ws, 5, 7)

    primeira = 6
    for i, row in agg.iterrows():
        r_ = primeira + i
        ws.cell(row=r_, column=2, value=row["codigo_tecnologia"])
        ws.cell(row=r_, column=3, value=float(row["potencia_total"]))
        ws.cell(row=r_, column=4, value=int(row["quantidade"]))
        # Representatividade = quantidade / total geral de quantidade (a fórmula referencia a linha do total)
        ws.cell(row=r_, column=5, value=f"=IFERROR(D{r_}/$D${primeira + len(agg) + 1},0)")
        ws.cell(row=r_, column=5).number_format = "0.0%"
        # Carga (kW) = Potência Total / 1000
        ws.cell(row=r_, column=6, value=f"=C{r_}/1000")
        # Consumo (kWh/mês) = Carga × (H + M/60) × 30
        ws.cell(row=r_, column=7, value=f"=F{r_}*($B$3+$C$3/60)*30")

    # Linha Total Geral
    ultima = primeira + len(agg) - 1
    linha_total = ultima + 2  # uma linha em branco entre a última e o total (padrão modelo)
    ws.cell(row=linha_total, column=2, value="Total Geral")
    ws.cell(row=linha_total, column=3, value=f"=SUM(C{primeira}:C{ultima})")
    ws.cell(row=linha_total, column=4, value=f"=SUM(D{primeira}:D{ultima})")
    ws.cell(row=linha_total, column=6, value=f"=SUM(F{primeira}:F{ultima})")
    ws.cell(row=linha_total, column=7, value=f"=SUM(G{primeira}:G{ultima})")
    aplicar_estilo_total(ws, linha_total, 7)

    autoajustar_largura(ws, 7, largura_minima=14)
    ws.freeze_panes = "A6"
