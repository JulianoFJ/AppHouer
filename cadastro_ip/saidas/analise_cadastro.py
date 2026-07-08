"""
Gerador do arquivo `Analise Cadastro.xlsx` (secao 12.2).

Estrutura (replicando o modelo ARMBH, com adaptacoes):
  - Aba `Cadastro` - cadastro normalizado
  - Aba `Base_HouerApp_IV` - inspecao bruta
  - Aba `Comparacao` - cruzamento ponto-a-ponto com flags de divergencia
  - Aba `Resultado Comparacao` - % de acerto (tecnologia/potencia/ambos)
  - Aba `Tratamento Convencional`
  - Aba `Tratamento LED IV` (com coluna Executado = Sim default)
  - Aba `Tratamento IAE`
  - Aba `Tratamento ID`
  - Aba `Resultado` - consolidado central com 3 blocos lado a lado
    (Cadastro Recebido | Cadastro Corrigido | Tratamento) + Tempo de Operacao,
    granularidade Tec x Pot, valores calculados em Python (sem SUMIFS frageis).
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ._helpers import (
    aplicar_estilo_header,
    aplicar_estilo_subheader,
    aplicar_estilo_total,
    autoajustar_largura,
    escrever_dataframe,
    gerar_combinacoes_tec_pot,
    workbook_para_bytes,
)
from ..reator import perda_reator


# Nomes das abas de tratamento
ABA_TRAT_CONV   = "Tratamento Convencional"
ABA_TRAT_LED_IV = "Tratamento LED IV"
ABA_TRAT_IAE    = "Tratamento IAE"
ABA_TRAT_ID     = "Tratamento ID"


def gerar(resultado) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _aba_cadastro(wb, resultado)
    _aba_base_houerapp_iv(wb, resultado)
    _aba_comparacao(wb, resultado)
    _aba_resultado_comparacao(wb, resultado)
    _aba_tratamento_convencional(wb, resultado)
    _aba_tratamento_led_iv(wb, resultado)
    _aba_tratamento_iae(wb, resultado)
    _aba_tratamento_id(wb, resultado)
    _aba_resultado(wb, resultado)

    return workbook_para_bytes(wb)


# ── Aba Cadastro ──────────────────────────────────────────────────────────────
def _aba_cadastro(wb: Workbook, r) -> None:
    ws = wb.create_sheet("Cadastro")
    df = r.cadastro_normalizado.copy()
    df = df.drop(columns=[c for c in ["familia_tecnologia"] if c in df.columns])
    escrever_dataframe(ws, df, linha_inicial=1)
    autoajustar_largura(ws, len(df.columns))
    ws.freeze_panes = "A2"


# ── Aba Base_HouerApp_IV ──────────────────────────────────────────────────────
def _aba_base_houerapp_iv(wb: Workbook, r) -> None:
    ws = wb.create_sheet("Base_HouerApp_IV")
    df = r.inspecao_normalizada.copy()
    df = df.drop(columns=[c for c in ["familia_tecnologia"] if c in df.columns])
    escrever_dataframe(ws, df, linha_inicial=1)
    autoajustar_largura(ws, len(df.columns))
    ws.freeze_panes = "A2"


# ── Aba Comparacao ────────────────────────────────────────────────────────────
def _aba_comparacao(wb: Workbook, r) -> None:
    ws = wb.create_sheet("Comparacao")
    df = r.comparacao.copy()
    df = df.drop(columns=[c for c in df.columns if "familia_tecnologia" in c.lower()], errors="ignore")
    escrever_dataframe(ws, df, linha_inicial=1)
    autoajustar_largura(ws, len(df.columns))
    ws.freeze_panes = "A2"


# ── Aba Resultado Comparacao ──────────────────────────────────────────────────
def _aba_resultado_comparacao(wb: Workbook, r) -> None:
    """Tabela dinamica de % de acerto usando valores calculados em Python."""
    ws = wb.create_sheet("Resultado Comparacao")
    ws["A1"] = "Metrica"
    ws["B1"] = "Acertos"
    ws["C1"] = "Total Inspecionados"
    ws["D1"] = "% de Acerto"
    aplicar_estilo_header(ws, 1, 4)

    comp = r.comparacao.copy() if r.comparacao is not None else pd.DataFrame()

    if comp.empty or "flag_tecnologia" not in comp.columns:
        ws["A2"] = "Sem pontos inspecionados para calcular acerto."
        autoajustar_largura(ws, 4, largura_minima=22)
        return

    # Filtra apenas pontos com inspecao. IMPORTANTE: o pipeline gera o valor
    # "Sem inspeção" (com cedilha e til). Comparar sem acento aqui faria a
    # máscara casar com ninguém e o total cair sobre o cadastro inteiro.
    insp_mask = comp["flag_tecnologia"] != "Sem inspeção"
    total_insp = int(insp_mask.sum())

    acerto_tec = int((comp.loc[insp_mask, "flag_tecnologia"] == "Igual").sum())
    acerto_pot = int((comp.loc[insp_mask, "flag_potencia"] == "Igual").sum()) if "flag_potencia" in comp.columns else 0
    acerto_ambos = int(
        ((comp.loc[insp_mask, "flag_tecnologia"] == "Igual") &
         (comp.loc[insp_mask, "flag_potencia"] == "Igual")).sum()
    ) if "flag_potencia" in comp.columns else acerto_tec

    pct_tec   = acerto_tec   / total_insp if total_insp > 0 else 0
    pct_pot   = acerto_pot   / total_insp if total_insp > 0 else 0
    pct_ambos = acerto_ambos / total_insp if total_insp > 0 else 0

    dados = [
        ("Tecnologia Correta",            acerto_tec,   total_insp, pct_tec),
        ("Potencia Correta",              acerto_pot,   total_insp, pct_pot),
        ("Tecnologia e Potencia Corretas", acerto_ambos, total_insp, pct_ambos),
    ]
    for i, (metrica, acerto, total, pct) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=metrica)
        ws.cell(row=i, column=2, value=acerto)
        ws.cell(row=i, column=3, value=total)
        ws.cell(row=i, column=4, value=pct)
        ws.cell(row=i, column=4).number_format = "0.0%"

    autoajustar_largura(ws, 4, largura_minima=22)


# ── Abas de Tratamento ────────────────────────────────────────────────────────
def _aba_tratamento_convencional(wb: Workbook, r) -> None:
    ws = wb.create_sheet(ABA_TRAT_CONV)
    df = r.tratamento_convencional.copy()
    cols = _selecionar_cols_trat_conv(df)
    df_out = df[cols].rename(columns=_renomear_trat_conv())
    escrever_dataframe(ws, df_out, linha_inicial=1)
    autoajustar_largura(ws, len(df_out.columns))
    ws.freeze_panes = "A2"


def _aba_tratamento_led_iv(wb: Workbook, r) -> None:
    ws = wb.create_sheet(ABA_TRAT_LED_IV)
    df = r.tratamento_led_iv.copy()
    cols = _selecionar_cols_trat_led_iv(df)
    df_out = df[cols].rename(columns=_renomear_trat_led_iv())
    escrever_dataframe(ws, df_out, linha_inicial=1)
    autoajustar_largura(ws, len(df_out.columns))
    ws.freeze_panes = "A2"


def _aba_tratamento_iae(wb: Workbook, r) -> None:
    ws = wb.create_sheet(ABA_TRAT_IAE)
    df = r.tratamento_iae.copy()
    if df.empty:
        ws["A1"] = "Sem pontos IAE para tratamento."
        return
    cols = _selecionar_cols_iae_id(df)
    df_out = df[cols].rename(columns=_renomear_iae_id())
    escrever_dataframe(ws, df_out, linha_inicial=1)
    autoajustar_largura(ws, len(df_out.columns))
    ws.freeze_panes = "A2"


def _aba_tratamento_id(wb: Workbook, r) -> None:
    ws = wb.create_sheet(ABA_TRAT_ID)
    df = r.tratamento_id.copy()
    if df.empty:
        ws["A1"] = "Sem pontos ID para tratamento."
        return
    cols = _selecionar_cols_iae_id(df)
    df_out = df[cols].rename(columns=_renomear_iae_id())
    escrever_dataframe(ws, df_out, linha_inicial=1)
    autoajustar_largura(ws, len(df_out.columns))
    ws.freeze_panes = "A2"


# ── Aba Resultado ─────────────────────────────────────────────────────────────
def _aba_resultado(wb: Workbook, r) -> None:
    """
    Estrutura (alinhada com o modelo ARMBH):
      Linha 1 - supertitulos: Cadastro Recebido | Cadastro Corrigido | Tratamento | Tempo de Operacao
      Linha 2 - subtitulos das colunas
      Linhas 3..N - dados, uma linha por combinacao Tec x Pot

    Blocos:
      A-E: Cadastro Recebido   -> Tecnologia | Potencia | Reator | Quantidade | Potencia Total
      F: separador
      G-K: Cadastro Corrigido  -> Tecnologia | Potencia | Reator | Quantidade | Potencia Total
      L: separador
      M-P: Tratamento          -> Convencional | LED | IAE | ID
      Q: separador
      R-S: Tempo de Operacao   -> Horas | Minutos

    USA resultado_por_tec_pot (pre-calculado em Python) — sem SUMIFS com letras
    de coluna fixas que variam por municipio e corrompem o arquivo .xlsx.
    """
    ws = wb.create_sheet("Resultado")

    # Cabecalhos de bloco (linha 1)
    ws["A1"] = "Cadastro Recebido"
    ws.merge_cells("A1:E1")
    ws["G1"] = "Cadastro Corrigido"
    ws.merge_cells("G1:K1")
    ws["M1"] = "Tratamento"
    ws.merge_cells("M1:P1")
    ws["R1"] = "Tempo de Operacao"
    ws.merge_cells("R1:S1")

    subheaders = {
        "A2": "Tecnologia", "B2": "Potencia (W)", "C2": "Reator (W)", "D2": "Quantidade", "E2": "Potencia Total (W)",
        "G2": "Tecnologia", "H2": "Potencia (W)", "I2": "Reator (W)", "J2": "Quantidade", "K2": "Potencia Total (W)",
        "M2": "Convencional", "N2": "LED", "O2": "IAE", "P2": "ID",
        "R2": "Horas", "S2": "Minutos",
    }
    for cel, valor in subheaders.items():
        ws[cel] = valor

    aplicar_estilo_header(ws, 1, 19)
    aplicar_estilo_subheader(ws, 2, 19)

    # Usa resultado_por_tec_pot (pre-calculado pelo pipeline)
    rtp = r.resultado_por_tec_pot if r.resultado_por_tec_pot is not None else pd.DataFrame()

    if rtp.empty:
        # Fallback: gera combos apenas do cadastro sem tratamentos
        combos = gerar_combinacoes_tec_pot(r.cadastro_normalizado, "codigo_tecnologia", "potencia")
        if combos.empty:
            ws["A3"] = "Sem dados para gerar Resultado."
            return
        rtp = combos.rename(columns={"codigo_tecnologia": "tec", "potencia": "pot", "quantidade": "qtd_recebida"})
        for c in ("qtd_corrigida", "trat_conv", "trat_led", "trat_iae", "trat_id"):
            if c not in rtp.columns:
                rtp[c] = 0
        rtp["qtd_corrigida"] = rtp["qtd_recebida"]

    primeira_dados = 3
    ultima_dados = primeira_dados + len(rtp) - 1

    # Escreve linhas de dados — usa enumerate(pos) para garantir linha correta
    for pos, (_, row) in enumerate(rtp.iterrows()):
        r_ = primeira_dados + pos

        # round() antes de int() — quantidades são discretas (lâmpadas), mas a
        # agregação por float acumula erros de precisão (ex: 1111 vira
        # 1110.9999999999998 e int() truncaria para 1110, perdendo 1 ponto por
        # bucket afetado).
        def _qtd(chave: str) -> int:
            v = row.get(chave, 0)
            return int(round(float(v))) if pd.notna(v) else 0

        tec     = str(row.get("tec", "") or "")
        pot     = float(row.get("pot", 0)) if pd.notna(row.get("pot")) else 0.0
        qtd_rec = _qtd("qtd_recebida")
        qtd_cor = _qtd("qtd_corrigida")
        tconv   = _qtd("trat_conv")
        tled    = _qtd("trat_led")
        tiae    = _qtd("trat_iae")
        tid     = _qtd("trat_id")

        # Reator pré-calculado em Python (igual para os dois blocos: mesmo Tec/Pot).
        # Substitui a fórmula =IF(IF(...)) que o Excel rejeitava como "fórmula
        # ilegível" e disparava recuperação ao abrir o arquivo.
        is_led = tec.upper() == "LED"
        reator_w = float(perda_reator(pot, is_led=is_led))

        # Bloco A-E — Cadastro Recebido
        ws.cell(row=r_, column=1, value=tec)
        ws.cell(row=r_, column=2, value=pot)
        ws.cell(row=r_, column=3, value=reator_w)
        ws.cell(row=r_, column=4, value=qtd_rec)
        ws.cell(row=r_, column=5, value=f"=(B{r_}+C{r_})*D{r_}")

        # Bloco G-K — Cadastro Corrigido (mesmo Tec/Pot, mesmo reator)
        ws.cell(row=r_, column=7,  value=tec)
        ws.cell(row=r_, column=8,  value=pot)
        ws.cell(row=r_, column=9,  value=reator_w)
        ws.cell(row=r_, column=10, value=qtd_cor)
        ws.cell(row=r_, column=11, value=f"=(H{r_}+I{r_})*J{r_}")

        # Bloco M-P — Tratamentos (valores diretos, sem SUMIFS)
        ws.cell(row=r_, column=13, value=tconv)
        ws.cell(row=r_, column=14, value=tled)
        ws.cell(row=r_, column=15, value=tiae)
        ws.cell(row=r_, column=16, value=tid)

    # Tempo de Operacao (cols R, S)
    if r.tempo_operacao is not None:
        ws.cell(row=3, column=18, value=int(r.tempo_operacao.horas))
        ws.cell(row=3, column=19, value=int(r.tempo_operacao.minutos))

    # Linha TOTAL
    linha_total = ultima_dados + 2
    ws.cell(row=linha_total, column=1, value="TOTAL")
    for col_letra in ["D", "E", "J", "K", "M", "N", "O", "P"]:
        col_idx = ord(col_letra) - 64
        ws.cell(
            row=linha_total,
            column=col_idx,
            value=f"=SUM({col_letra}{primeira_dados}:{col_letra}{ultima_dados})",
        )
    aplicar_estilo_total(ws, linha_total, 19)

    # Larguras fixas
    larguras = {
        "A": 12, "B": 13, "C": 11, "D": 12, "E": 16,
        "F": 2,
        "G": 12, "H": 13, "I": 11, "J": 12, "K": 16,
        "L": 2,
        "M": 14, "N": 12, "O": 12, "P": 12,
        "Q": 2,
        "R": 8, "S": 10,
    }
    for letra, largura in larguras.items():
        ws.column_dimensions[letra].width = largura

    ws.freeze_panes = "A3"


# ── Helpers de selecao/renomeacao de colunas ──────────────────────────────────
def _letra_da_coluna(cols: list[str], nome: str) -> str:
    """Retorna a letra Excel correspondente a coluna `nome` na lista dada."""
    try:
        idx = cols.index(nome)
    except ValueError:
        return "A"
    return get_column_letter(idx + 1)


def _selecionar_cols_trat_conv(df: pd.DataFrame) -> list[str]:
    preferenciais = [
        "id_ponto", "logradouro", "codigo_tecnologia", "potencia",
        "tecnologia_inspecao", "potencia_inspecao",
        "quantidade_extrapolada", "reator_w",
        "fator_extrapolacao", "latitude", "longitude",
    ]
    return [c for c in preferenciais if c in df.columns]


def _selecionar_cols_trat_led_iv(df: pd.DataFrame) -> list[str]:
    preferenciais = [
        "id_ponto", "logradouro", "codigo_tecnologia", "potencia",
        "tecnologia_inspecao", "potencia_inspecao", "quantidade_inspecao",
        "reator_w", "fator_extrapolacao", "quantidade_considerada",
        "executado", "latitude", "longitude",
    ]
    return [c for c in preferenciais if c in df.columns]


def _selecionar_cols_iae_id(df: pd.DataFrame) -> list[str]:
    preferenciais = [
        "id_ponto", "logradouro", "bairro", "local",
        "codigo_tecnologia", "potencia", "quantidade",
        "reator_w", "quantidade_considerada",
        "origem", "latitude", "longitude",
    ]
    return [c for c in preferenciais if c in df.columns]


def _renomear_trat_conv() -> dict[str, str]:
    return {
        "id_ponto": "ID do Ponto",
        "logradouro": "Logradouro",
        "codigo_tecnologia": "Tec Cadastro",
        "potencia": "Pot Cadastro (W)",
        "tecnologia_inspecao": "Tec Inspecao",
        "potencia_inspecao": "Pot Inspecao (W)",
        "quantidade_extrapolada": "Quantidade de Pontos",
        "reator_w": "Reator (W)",
        "fator_extrapolacao": "Fator",
        "latitude": "Latitude",
        "longitude": "Longitude",
    }


def _renomear_trat_led_iv() -> dict[str, str]:
    return {
        "id_ponto": "ID do Ponto",
        "logradouro": "Logradouro",
        "codigo_tecnologia": "Tec Cadastro",
        "potencia": "Pot Cadastro (W)",
        "tecnologia_inspecao": "Tec Inspecao",
        "potencia_inspecao": "Pot Inspecao (W)",
        "quantidade_inspecao": "Qtd Inspecao",
        "reator_w": "Reator (W)",
        "fator_extrapolacao": "Fator",
        "quantidade_considerada": "Quantidade Considerada",
        "executado": "Executado",
        "latitude": "Latitude",
        "longitude": "Longitude",
    }


def _renomear_iae_id() -> dict[str, str]:
    return {
        "id_ponto": "ID do Ponto",
        "logradouro": "Logradouro",
        "bairro": "Bairro",
        "local": "Tipo de Local",
        "codigo_tecnologia": "Tecnologia",
        "potencia": "Potencia (W)",
        "quantidade": "Quantidade",
        "reator_w": "Reator (W)",
        "quantidade_considerada": "Quantidade Considerada",
        "origem": "Origem",
        "latitude": "Latitude",
        "longitude": "Longitude",
    }
