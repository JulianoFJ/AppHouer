"""
Helpers comuns para os 3 geradores de saída.

Estilização padronizada (paleta Navy + Teal), formatação numérica,
e utilidades para gravar DataFrame + fórmulas vivas usando openpyxl direto
(não via pandas.to_excel, que não preserva fórmulas como strings).
"""

from __future__ import annotations

from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── Paleta da marca ──────────────────────────────────────────────────────────────
PALETA_NAVY = "1B3664"
PALETA_TEAL = "00A9E0"
LIGHT_GRAY = "F1F5F9"
BORDER_GRAY = "94A3B8"


# ── Estilos pré-fabricados ────────────────────────────────────────────────────
def header_fill() -> PatternFill:
    return PatternFill(start_color=PALETA_NAVY, end_color=PALETA_NAVY, fill_type="solid")


def subheader_fill() -> PatternFill:
    return PatternFill(start_color=PALETA_TEAL, end_color=PALETA_TEAL, fill_type="solid")


def total_fill() -> PatternFill:
    return PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


def header_font() -> Font:
    return Font(name="Calibri", size=11, bold=True, color="FFFFFF")


def body_font() -> Font:
    return Font(name="Calibri", size=10, color="000000")


def total_font() -> Font:
    return Font(name="Calibri", size=10, bold=True, color="000000")


def thin_border() -> Border:
    side = Side(style="thin", color=BORDER_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)


def center_align() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


# ── Utilidades ────────────────────────────────────────────────────────────────
def autoajustar_largura(ws: Worksheet, colunas: int, largura_minima: int = 12, largura_maxima: int = 40) -> None:
    """Define largura por coluna olhando o conteúdo da linha 1 (cabeçalho)."""
    for c in range(1, colunas + 1):
        valor = ws.cell(row=1, column=c).value
        nome = "" if valor is None else str(valor)
        largura = max(largura_minima, min(largura_maxima, len(nome) + 4))
        ws.column_dimensions[get_column_letter(c)].width = largura


def aplicar_estilo_header(ws: Worksheet, linha: int, colunas: int) -> None:
    fill = header_fill()
    font = header_font()
    align = center_align()
    border = thin_border()
    for c in range(1, colunas + 1):
        cel = ws.cell(row=linha, column=c)
        cel.fill = fill
        cel.font = font
        cel.alignment = align
        cel.border = border


def aplicar_estilo_subheader(
    ws: Worksheet,
    linha: int,
    colunas: int,
    colunas_a_estilizar: list[int] | None = None,
) -> None:
    """
    Estiliza células de subheader (cabeçalho secundário, fundo teal).

    Args:
        linha: índice da linha (1-based)
        colunas: extensão padrão (1..colunas) quando `colunas_a_estilizar` é None
        colunas_a_estilizar: se fornecido, estiliza apenas estas colunas específicas
            — útil quando o subheader não ocupa todas as colunas adjacentes
            (ex: bloco isolado em H/I sem afetar A-E que tem dados na mesma linha).
    """
    fill = subheader_fill()
    font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    align = center_align()
    border = thin_border()
    indices = colunas_a_estilizar if colunas_a_estilizar is not None else list(range(1, colunas + 1))
    for c in indices:
        cel = ws.cell(row=linha, column=c)
        cel.fill = fill
        cel.font = font
        cel.alignment = align
        cel.border = border


def aplicar_estilo_total(ws: Worksheet, linha: int, colunas: int) -> None:
    fill = total_fill()
    font = total_font()
    border = thin_border()
    for c in range(1, colunas + 1):
        cel = ws.cell(row=linha, column=c)
        cel.fill = fill
        cel.font = font
        cel.border = border


def escrever_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    linha_inicial: int = 1,
    coluna_inicial: int = 1,
    com_cabecalho: bool = True,
    estilo_header: bool = True,
) -> int:
    """
    Escreve um DataFrame em uma worksheet preservando fórmulas (strings que começam com '=').

    Returns:
        Próxima linha vazia após o DataFrame (útil para encadear escritas).
    """
    linha = linha_inicial
    if com_cabecalho:
        for j, col in enumerate(df.columns):
            ws.cell(row=linha, column=coluna_inicial + j, value=str(col))
        if estilo_header:
            aplicar_estilo_header(ws, linha, coluna_inicial - 1 + len(df.columns))
        linha += 1

    for _, row in df.iterrows():
        for j, val in enumerate(row.values):
            v = _converter_valor(val)
            ws.cell(row=linha, column=coluna_inicial + j, value=v)
        linha += 1
    return linha


def escrever_formulas_coluna(
    ws: Worksheet,
    coluna: int,
    primeira_linha: int,
    ultima_linha: int,
    template: str,
) -> None:
    """
    Escreve uma fórmula viva em um intervalo de linhas, substituindo `{r}` pelo número da linha.

    Exemplo:
        escrever_formulas_coluna(ws, 5, 3, 100, "=(B{r}+C{r})*D{r}")
        → grava em E3..E100 a fórmula correspondente.
    """
    for r in range(primeira_linha, ultima_linha + 1):
        ws.cell(row=r, column=coluna, value=template.replace("{r}", str(r)))


def _converter_valor(v):
    """Converte valor do pandas para algo aceito pelo openpyxl."""
    if v is None:
        return None
    if pd.isna(v):
        return None
    # Strings que começam com '=' são gravadas como fórmula viva
    if isinstance(v, str):
        return v
    # int/float/bool são aceitos diretamente
    if isinstance(v, (int, float, bool)):
        return v
    # Datetime
    if hasattr(v, "isoformat"):
        return v
    # Outros (Timestamp, etc.) — converter para string
    return str(v)


def gerar_combinacoes_tec_pot(df: pd.DataFrame, col_tec: str = "codigo_tecnologia", col_pot: str = "potencia") -> pd.DataFrame:
    """
    Retorna DataFrame com combinações únicas Tecnologia × Potência,
    ordenadas por tecnologia e potência crescente.

    Granularidade exigida pelas instruções v1.4 (seções 12.2.1 e 12.3):
    uma linha por combinação Tec × Pot, agrupadas por tecnologia.
    """
    if col_tec not in df.columns or col_pot not in df.columns:
        return pd.DataFrame(columns=[col_tec, col_pot, "quantidade"])

    df2 = df.copy()
    df2[col_pot] = pd.to_numeric(df2[col_pot], errors="coerce")
    df2 = df2.dropna(subset=[col_tec])

    qtd_col = "quantidade" if "quantidade" in df2.columns else "Quantidade" if "Quantidade" in df2.columns else None
    if qtd_col:
        df2[qtd_col] = pd.to_numeric(df2[qtd_col], errors="coerce").fillna(1)
        grouped = df2.groupby([col_tec, col_pot], dropna=False)[qtd_col].sum().reset_index()
        grouped.columns = [col_tec, col_pot, "quantidade"]
    else:
        grouped = df2.groupby([col_tec, col_pot], dropna=False).size().reset_index(name="quantidade")

    # Ordenação: agrupado por tecnologia, potência crescente dentro de cada
    grouped["_ord_tec"] = grouped[col_tec].map(_ordem_tecnologia)
    grouped = grouped.sort_values(["_ord_tec", col_tec, col_pot], na_position="last").drop(columns=["_ord_tec"])
    grouped = grouped.reset_index(drop=True)
    return grouped


_ORDEM_PADRAO = ["LED", "VS", "VM", "VMT", "MT", "FL", "IN", "GLOBO", "ORNAMENTAL", "REBATEDOR", "PROJETOR"]


def _ordem_tecnologia(t) -> int:
    if t is None or (isinstance(t, float) and pd.isna(t)):
        return 99
    try:
        return _ORDEM_PADRAO.index(str(t).upper())
    except ValueError:
        return 50


def workbook_para_bytes(wb: Workbook) -> bytes:
    """Salva o workbook em memória e retorna os bytes do arquivo .xlsx."""
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = [
    "PALETA_NAVY", "PALETA_TEAL",
    "header_fill", "subheader_fill", "total_fill",
    "aplicar_estilo_header", "aplicar_estilo_subheader", "aplicar_estilo_total",
    "escrever_dataframe", "escrever_formulas_coluna",
    "gerar_combinacoes_tec_pot", "autoajustar_largura",
    "workbook_para_bytes",
]
