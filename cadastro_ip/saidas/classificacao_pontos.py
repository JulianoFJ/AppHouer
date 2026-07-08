"""
Gerador do arquivo `Classificação dos Pontos.xlsx` (seção 12.1).

Estrutura (replicando o modelo ARMBH):
  - Aba `Cadastro` — cadastro completo normalizado + coluna Classe propagada
  - Aba `Amostra` — pontos inspecionados com classes viária e de pedestre
  - Abas por classe encontrada (`1. C0`, `2. C1`, `3. C2`, `4. C3`, etc.) —
    cada uma lista os pontos do cadastro que receberam aquela classe
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from ._helpers import (
    aplicar_estilo_header,
    autoajustar_largura,
    escrever_dataframe,
    workbook_para_bytes,
)


# Ordem das classes a serem materializadas em subabas (apenas se houver pontos)
ORDEM_CLASSES = ["C0", "C1", "C2", "C3", "C4", "C5", "M1", "M2", "M3", "M4", "M5", "M6"]


def gerar(resultado) -> bytes:
    """
    Args:
        resultado: instância de `cadastro_ip.pipeline.ResultadoPipeline`.

    Returns:
        Bytes do arquivo .xlsx.
    """
    wb = Workbook()
    # Remove a worksheet default vazia
    wb.remove(wb.active)

    cad = resultado.propagacao_classe.df_cadastro_com_classe.copy()

    # ── Aba Cadastro ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Cadastro")
    cols_cadastro = _selecionar_colunas_cadastro(cad)
    df_cadastro = cad[cols_cadastro].copy()
    df_cadastro = df_cadastro.rename(columns=_renomear_para_exibicao(cols_cadastro))
    escrever_dataframe(ws, df_cadastro, linha_inicial=1)
    autoajustar_largura(ws, len(df_cadastro.columns))
    ws.freeze_panes = "A2"

    # ── Aba Amostra ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Amostra")
    insp = resultado.inspecao_normalizada.copy()
    cols_amostra = _selecionar_colunas_amostra(insp)
    df_amostra = insp[cols_amostra].rename(columns=_renomear_para_exibicao(cols_amostra))
    escrever_dataframe(ws, df_amostra, linha_inicial=1)
    autoajustar_largura(ws, len(df_amostra.columns))
    ws.freeze_panes = "A2"

    # ── Subabas por classe ──────────────────────────────────────────────────
    classes_presentes = [c for c in ORDEM_CLASSES if c in df_cadastro.get("Classe", pd.Series([], dtype=object)).unique()]
    for idx, classe in enumerate(classes_presentes, 1):
        nome_aba = f"{idx}. {classe}"
        ws = wb.create_sheet(nome_aba)
        subset = df_cadastro[df_cadastro["Classe"] == classe].copy()
        escrever_dataframe(ws, subset, linha_inicial=1)
        autoajustar_largura(ws, len(subset.columns))
        ws.freeze_panes = "A2"

    return workbook_para_bytes(wb)


def _selecionar_colunas_cadastro(cad: pd.DataFrame) -> list[str]:
    """Seleciona as colunas do cadastro normalizado que devem aparecer no arquivo de saída."""
    preferenciais = [
        "id_ponto", "sequencia", "municipio", "nome_cidade", "posicao", "endereco",
        "logradouro", "bairro", "local", "localizacao",
        "tecnologia", "tipo_lampada", "codigo_tecnologia",
        "potencia", "Quantidade", "quantidade",
        "tipo_luminaria", "material",
        "latitude", "longitude",
        "classe_via",
    ]
    return [c for c in preferenciais if c in cad.columns]


def _selecionar_colunas_amostra(insp: pd.DataFrame) -> list[str]:
    """Seleciona colunas da inspeção que devem aparecer na aba Amostra."""
    preferenciais = [
        "id_ponto", "identificador",
        "tipoInspecao", "tipo_inspecao",
        "logradouro",
        "latitude", "longitude",
        "tecnologia", "tipoLampada", "codigo_tecnologia",
        "potencia", "potenciaLampada",
        "quantidade", "quantidadeLampadas",
        "classe_via", "classeIluminacao",
        "classe_pedestre",
        "Latitude Cad.", "Longitude Cad.", "Realocado",
    ]
    return [c for c in preferenciais if c in insp.columns]


def _renomear_para_exibicao(cols: list[str]) -> dict[str, str]:
    """Mapeia nomes internos canônicos para os rótulos exibidos no .xlsx."""
    return {
        "id_ponto": "ID do Ponto",
        "sequencia": "Sequência",
        "nome_cidade": "Município",
        "municipio": "Município",
        "posicao": "Posição",
        "endereco": "Endereço",
        "logradouro": "Logradouro",
        "bairro": "Bairro",
        "localizacao": "Localização",
        "local": "Tipo de Local",
        "tecnologia": "Tecnologia",
        "tipo_lampada": "Tipo Lâmpada",
        "codigo_tecnologia": "Código Tec.",
        "potencia": "Potência (W)",
        "potenciaLampada": "Potência (W)",
        "Quantidade": "Quantidade",
        "quantidade": "Quantidade",
        "quantidadeLampadas": "Quantidade",
        "tipo_luminaria": "Tipo Luminária",
        "material": "Material",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "classe_via": "Classe",
        "classeIluminacao": "Classe Inspeção",
        "classe_pedestre": "Classe Pedestre",
        "tipoInspecao": "Tipo Inspeção",
        "tipoLampada": "Tipo Lâmpada",
    }
