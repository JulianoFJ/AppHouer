"""
Monta o cadastro tratado que a página de amostragem consome e que o usuário baixa.

Junta as três peças — o cadastro extraído da BDGD (`extracao`), o logradouro e a classe
do OpenStreetMap (`vias_osm`) — num DataFrame com nomes de coluna que o detector de
`cadastro_ip.normalizacao` reconhece sozinho. É o que permite a página tratar a origem
BDGD exatamente como trata uma planilha enviada: a partir daqui o fluxo é o mesmo.

A procedência anda junto com o dado
-----------------------------------
Um cadastro montado de três fontes, com duas colunas inferidas por proximidade e uma
classe estimada por método normativo incompleto, não pode circular sem dizer isso. A
planilha de download tem uma aba **Procedência** com a origem de cada coluna e as
ressalvas — quem receber o arquivo dois meses depois precisa saber o que é medido e o
que é inferido, sem ter de perguntar.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cadastro_ip.saidas._helpers import (
    aplicar_estilo_header,
    autoajustar_largura,
    escrever_dataframe,
    header_font,
    subheader_fill,
    workbook_para_bytes,
)

from . import extracao, vias_osm

# Ordem de exibição. As seis primeiras são as que a amostragem procura pelo nome; o
# resto é contexto de engenharia que a equipe de campo usa para conferir o ponto.
COLUNAS_SAIDA = [
    "id_ponto", "logradouro", "classe_via", "latitude", "longitude",
    "tecnologia", "potencia_lampada_w", "carga_instalada_kw",
    "hierarquia_osm", "dist_via_m", "area_urbana", "telegestao",
    "tipo_sistema", "situacao", "data_conexao",
    "municipio_ibge", "distribuidora", "data_base_bdgd", "versao_bdgd",
    "metodo_classe",
]

PROCEDENCIA = [
    ("id_ponto", "BDGD / PIP", "COD_ID da unidade consumidora de iluminação pública. Medido."),
    ("latitude, longitude", "BDGD / PONNOT", "Coordenada do ponto notável de conexão (PN_CON), "
     "SIRGAS 2000. Medido — mas é a posição do poste da rede, não do luminário."),
    ("logradouro", "OpenStreetMap", "Nome da via mais próxima do ponto. INFERIDO por proximidade."),
    ("classe_via", "NBR 5101:2024 + OSM", "Classe M estimada pela Tabela 1 da norma, com os "
     "parâmetros disponíveis. ESTIMADO — não é enquadramento normativo."),
    ("hierarquia_osm", "OpenStreetMap", "Hierarquia funcional da via casada. INFERIDO."),
    ("dist_via_m", "calculado", "Distância do ponto ao eixo da via casada. Use para julgar a "
     "confiança do logradouro e da classe."),
    ("tecnologia", "BDGD / PIP", "INFERIDA da assinatura física do código TIPO_LAMP (perda de "
     "reator e série de potências) — a BDGD não publica o domínio do campo."),
    ("potencia_lampada_w", "BDGD / PIP", "POT_LAMP. Declarado pela distribuidora."),
    ("carga_instalada_kw", "BDGD / PIP", "CAR_INST. Declarado pela distribuidora."),
    ("area_urbana", "BDGD / PIP", "ARE_LOC = UB. Declarado."),
    ("telegestao", "BDGD / PIP", "CONTROLE > 0. Declarado — conferir em campo, o preenchimento "
     "deste campo é irregular entre distribuidoras."),
]


@dataclass
class CadastroMontado:
    dados: pd.DataFrame
    codigo_ibge: str
    ressalvas: list[str] = field(default_factory=list)
    enriquecido: bool = False

    @property
    def total(self) -> int:
        return len(self.dados)

    @property
    def com_coordenada(self) -> int:
        return int(self.dados["latitude"].notna().sum())

    @property
    def com_logradouro(self) -> int:
        if "logradouro" not in self.dados.columns:
            return 0
        return int((self.dados["logradouro"].fillna("") != "").sum())

    @property
    def distribuidora(self) -> str:
        if self.dados.empty or "distribuidora" not in self.dados.columns:
            return ""
        return str(self.dados["distribuidora"].iloc[0])

    @property
    def data_base(self) -> str:
        if self.dados.empty or "data_base_bdgd" not in self.dados.columns:
            return ""
        return str(self.dados["data_base_bdgd"].iloc[0])


def disponivel(codigo_ibge: str) -> bool:
    """Há cadastro extraído para este município? A página usa isto para decidir a UI."""
    from . import caminhos
    return caminhos.caminho_cadastro(str(codigo_ibge)) is not None


def listar_disponiveis() -> pd.DataFrame:
    """
    Os municípios com cadastro já extraído, com nome e UF resolvidos.

    Colunas: `codigo_ibge`, `ente`, `uf`, `rotulo`, `pontos`, `kb`, `publicado`.

    A página lista **só isto** na origem BDGD, em vez dos 5.570 do IBGE: o cadastro
    ponto a ponto exige o `.gdb` da distribuidora e GDAL, que não existem no Streamlit
    Cloud. Oferecer um município que não vai abrir seria pior que não oferecer.
    """
    from hub_municipios import config

    from . import caminhos

    codigos = caminhos.municipios_disponiveis()
    if not codigos:
        return pd.DataFrame(columns=["codigo_ibge", "ente", "uf", "rotulo",
                                     "pontos", "kb", "publicado"])

    try:
        entes = pd.read_parquet(config.ENTES_CACHE, columns=["cod_ibge", "ente", "uf"])
        entes["cod_ibge"] = entes["cod_ibge"].astype(str)
    except Exception:                                   # noqa: BLE001
        entes = pd.DataFrame(columns=["cod_ibge", "ente", "uf"])

    linhas = []
    for codigo in codigos:
        arquivo = caminhos.caminho_cadastro(codigo)
        if arquivo is None:
            continue
        achado = entes[entes["cod_ibge"] == codigo]
        nome = str(achado.iloc[0]["ente"]) if not achado.empty else codigo
        uf = str(achado.iloc[0]["uf"]) if not achado.empty else ""
        try:
            pontos = len(pd.read_parquet(arquivo, columns=["id_ponto"]))
        except Exception:                               # noqa: BLE001
            pontos = 0
        linhas.append({
            "codigo_ibge": codigo,
            "ente": nome,
            "uf": uf,
            "rotulo": f"{nome}/{uf}" if uf else nome,
            "pontos": pontos,
            "kb": round(arquivo.stat().st_size / 1e3),
            "publicado": caminhos.esta_publicado(codigo),
        })
    return pd.DataFrame(linhas).sort_values("rotulo").reset_index(drop=True)


def osm_em_cache(codigo_ibge: str) -> bool:
    """
    A malha viária deste município já está em disco?

    A página usa isto para decidir se mostra botão. A regra do projeto é não pedir
    clique quando não há rede a pagar: com o parquet extraído e o OSM em cache, montar
    o cadastro é trabalho local e pode acontecer direto.
    """
    from . import caminhos
    return (caminhos.OSM_CACHE / f"{codigo_ibge}.json").exists()


def montar(codigo_ibge: str, enriquecer: bool = True,
           usar_cache_osm: bool = True) -> Optional[CadastroMontado]:
    """
    Lê o cadastro extraído e, opcionalmente, enriquece com OSM. None se não existir.

    O enriquecimento é a única parte que toca a rede. Quando ele falha — Overpass fora
    do ar, sem internet — o cadastro **continua sendo devolvido** sem logradouro e sem
    classe, com a ressalva explicando: perder a amostra inteira porque um serviço
    público de cortesia está indisponível seria pior que sortear com menos estratos.
    """
    cru = extracao.carregar(codigo_ibge)
    if cru is None:
        return None

    ressalvas: list[str] = []
    enriquecido = False
    dados = cru

    if enriquecer:
        try:
            dados, ressalvas_osm = vias_osm.enriquecer(cru, str(codigo_ibge),
                                                       usar_cache=usar_cache_osm)
            ressalvas.extend(ressalvas_osm)
            enriquecido = True
        except Exception as exc:                        # noqa: BLE001
            ressalvas.append(
                f"O enriquecimento pelo OpenStreetMap falhou ({exc}). O cadastro saiu "
                "sem logradouro e sem classe viária: o sorteio perde as cotas por via "
                "estruturante e por classe, e fica com dispersão geográfica e "
                "estratificação por tecnologia."
            )
    else:
        ressalvas.append(
            "Enriquecimento pelo OpenStreetMap desligado: sem logradouro e sem classe "
            "viária, o sorteio não garante cobertura de avenida e rodovia."
        )

    sem_coordenada = int(dados["latitude"].isna().sum())
    if sem_coordenada:
        ressalvas.insert(0, (
            f"{sem_coordenada:,} pontos ficaram sem coordenada porque o PN_CON não casou "
            "com nenhum ponto notável da rede. Eles continuam no sorteio, mas fora da "
            "dispersão espacial e do mapa.".replace(",", ".")
        ))

    ressalvas.insert(0, (
        "Cadastro derivado da BDGD da ANEEL, não do cadastro municipal. Ele cobre o que "
        "a distribuidora fatura como iluminação pública — ponto não faturado, ou faturado "
        "em outro município, não aparece."
    ))

    ordenadas = [c for c in COLUNAS_SAIDA if c in dados.columns]
    restantes = [c for c in dados.columns if c not in ordenadas]
    return CadastroMontado(dados=dados[ordenadas + restantes],
                           codigo_ibge=str(codigo_ibge),
                           ressalvas=ressalvas, enriquecido=enriquecido)


# ── Planilha para download ───────────────────────────────────────────────────

def _aba_cadastro(wb: Workbook, montado: CadastroMontado) -> None:
    ws = wb.active
    ws.title = "Cadastro"
    escrever_dataframe(ws, montado.dados, linha_inicial=1)
    aplicar_estilo_header(ws, 1, len(montado.dados.columns))
    ws.freeze_panes = "B2"
    autoajustar_largura(ws, len(montado.dados.columns))


def _aba_procedencia(wb: Workbook, montado: CadastroMontado) -> None:
    ws = wb.create_sheet("Procedência")
    ws["A1"] = f"Cadastro de IP — município {montado.codigo_ibge}"
    ws["A1"].font = header_font()
    ws["A1"].fill = subheader_fill()

    linhas = [
        ("Origem", "BDGD (ANEEL), entidade PIP + PONNOT"),
        ("Distribuidora", montado.distribuidora),
        ("Data-base da BDGD", montado.data_base),
        ("Pontos", f"{montado.total:,}".replace(",", ".")),
        ("Com coordenada", f"{montado.com_coordenada:,}".replace(",", ".")),
        ("Com logradouro", f"{montado.com_logradouro:,}".replace(",", ".")),
        ("Enriquecido com OSM", "sim" if montado.enriquecido else "não"),
    ]
    linha = 3
    for rotulo, valor in linhas:
        ws.cell(row=linha, column=1, value=rotulo)
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="Coluna").font = header_font()
    ws.cell(row=linha, column=2, value="Fonte").font = header_font()
    ws.cell(row=linha, column=3, value="Observação").font = header_font()
    aplicar_estilo_header(ws, linha, 3)
    linha += 1
    for coluna, fonte, nota in PROCEDENCIA:
        ws.cell(row=linha, column=1, value=coluna)
        ws.cell(row=linha, column=2, value=fonte)
        ws.cell(row=linha, column=3, value=nota)
        linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="Ressalvas").font = header_font()
    linha += 1
    for ressalva in montado.ressalvas:
        ws.cell(row=linha, column=1, value="•")
        ws.cell(row=linha, column=2, value=ressalva)
        linha += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 88


def planilha(montado: CadastroMontado) -> bytes:
    """Cadastro tratado em .xlsx, com a aba de procedência junto."""
    wb = Workbook()
    _aba_cadastro(wb, montado)
    _aba_procedencia(wb, montado)
    return workbook_para_bytes(wb)


def nome_arquivo(montado: CadastroMontado) -> str:
    return f"cadastro_bdgd_{montado.codigo_ibge}_{montado.data_base or 'sem-data'}.xlsx"


__all__ = ["COLUNAS_SAIDA", "PROCEDENCIA", "CadastroMontado", "disponivel",
           "montar", "planilha", "nome_arquivo"]
