"""
Geração das duas planilhas de campo — estrutural e qualidade.

Cada arquivo sai com três abas, na ordem em que a equipe usa:

  1. **Amostra de Campo** — o formulário. Traz as colunas de identificação e
     localização do ponto sorteado seguidas das colunas em branco a preencher em
     campo, já com validação de lista onde faz sentido. É a aba que vai para o tablet.
  2. **Cadastro (referência)** — as mesmas linhas sorteadas com todas as colunas
     originais do cadastro municipal, para conferência do que estava declarado.
  3. **Plano de Amostragem** — a memória de cálculo: parque, plano NBR 5426, semente
     do sorteio, cobertura por classe e vias principais contempladas. É a aba que
     sustenta o dado perante o poder concedente e a banca.

Reaproveita a estilização do portal (`cadastro_ip/saidas/_helpers.py`) para que as
planilhas de amostragem tenham a mesma cara das três saídas da Análise de Cadastro.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cadastro_ip.saidas._helpers import (
    aplicar_estilo_header,
    autoajustar_largura,
    escrever_dataframe,
    header_font,
    subheader_fill,
    workbook_para_bytes,
)

from ..amostrador import COLUNAS_AUXILIARES, GRUPO_ESTRUTURAL, GRUPO_QUALIDADE, ResultadoAmostragem


# ── Colunas de preenchimento em campo ─────────────────────────────────────────
# (rótulo, lista de opções para validação — None = texto/número livre)
CAMPOS_ESTRUTURAL: list[tuple[str, list[str] | None]] = [
    ("Data da inspeção", None),
    ("Equipe / responsável", None),
    ("Ponto localizado?", ["Sim", "Não — inexistente", "Não — inacessível"]),
    ("Latitude aferida", None),
    ("Longitude aferida", None),
    ("Poste — material", ["Concreto", "Metálico", "Madeira", "Fibra", "Outro"]),
    ("Poste — altura livre (m)", None),
    ("Poste — estado", ["Bom", "Regular", "Ruim", "Substituir"]),
    ("Braço — comprimento (m)", None),
    ("Braço — inclinação (°)", None),
    ("Altura de montagem (m)", None),
    ("Luminária — tipo/modelo", None),
    ("Luminária — tecnologia", ["LED", "VS", "VM", "MT", "FL", "IN", "Outra"]),
    ("Luminária — potência (W)", None),
    ("Luminária — estado", ["Bom", "Regular", "Ruim", "Substituir"]),
    ("Qtd. de luminárias no poste", None),
    ("Rede", ["Aérea", "Subterrânea", "Mista"]),
    ("Fotocélula / relé", ["Individual", "Comando em grupo", "Telegestão", "Ausente"]),
    ("Via — largura da pista (m)", None),
    ("Via — nº de faixas", None),
    ("Via — largura do passeio (m)", None),
    ("Via — afastamento do poste (m)", None),
    ("Disposição das luminárias", ["Unilateral", "Bilateral alternada", "Bilateral frente a frente", "Canteiro central", "Outra"]),
    ("Vão entre postes (m)", None),
    ("Arborização interferindo", ["Não", "Leve", "Moderada", "Intensa"]),
    ("Classe viária confirmada?", ["Sim", "Não — corrigir"]),
    ("Classe viária aferida em campo", None),
    ("Divergência com o cadastro", ["Não", "Sim — tecnologia", "Sim — potência", "Sim — localização", "Sim — outra"]),
    ("Nº da foto", None),
    ("Observações", None),
]

CAMPOS_QUALIDADE: list[tuple[str, list[str] | None]] = [
    ("Data da medição", None),
    ("Hora início", None),
    ("Hora fim", None),
    ("Equipe / responsável", None),
    ("Luxímetro — modelo", None),
    ("Luxímetro — nº de série", None),
    ("Certificado de calibração", None),
    ("Condição do tempo", ["Seco", "Chuva", "Neblina", "Pós-chuva"]),
    ("Trecho medido — do poste", None),
    ("Trecho medido — ao poste", None),
    ("Vão entre postes (m)", None),
    ("Altura de montagem (m)", None),
    ("Largura da pista (m)", None),
    ("Nº de faixas", None),
    ("Disposição das luminárias", ["Unilateral", "Bilateral alternada", "Bilateral frente a frente", "Canteiro central", "Outra"]),
    ("Malha — nº de linhas", None),
    ("Malha — nº de colunas", None),
    ("Total de pontos medidos", None),
    ("E mínima medida (lux)", None),
    ("E máxima medida (lux)", None),
    ("E média medida (lux)", None),
    ("Uniformidade U0 (Emín/Eméd)", None),
    ("Uniformidade longitudinal Ul", None),
    ("Classe de iluminação adotada", None),
    ("E média requerida pela NBR 5101 (lux)", None),
    ("U0 requerida pela NBR 5101", None),
    ("Atende à NBR 5101?", ["Sim", "Não — iluminância", "Não — uniformidade", "Não — ambos"]),
    ("Luminárias apagadas no trecho", None),
    ("Nº da foto", None),
    ("Observações", None),
]

ROTULO_GRUPO = {
    GRUPO_ESTRUTURAL: "Medição Estrutural",
    GRUPO_QUALIDADE: "Medição de Qualidade",
}

# Colunas de identificação que abrem o formulário de campo.
_COLUNAS_IDENTIFICACAO = [
    ("Nº", "_ordem"),
    ("ID do ponto", "_id"),
    ("Logradouro", "_logradouro"),
    ("Bairro", "_bairro"),
    ("Classe (cadastro)", "_classe"),
    ("Tipo de via", "_tipo_via"),
    ("Via principal", "_via_principal"),
    ("Latitude (cadastro)", "_lat"),
    ("Longitude (cadastro)", "_lon"),
]


def _formulario(
    resultado: ResultadoAmostragem, grupo: str
) -> tuple[pd.DataFrame, list[tuple[str, list[str] | None]]]:
    """Monta o DataFrame do formulário de campo (identificação + colunas em branco)."""
    amostra = (resultado.estrutural if grupo == GRUPO_ESTRUTURAL else resultado.qualidade).copy()
    chaves_principais = {v.chave for v in resultado.vias_principais}
    amostra["_via_principal"] = amostra["_chave_via"].map(
        lambda c: "Sim" if c in chaves_principais else "Não"
    )
    # Ordena por bairro e logradouro: é como a equipe percorre o município, e reduz
    # deslocamento entre pontos — a aleatoriedade já foi decidida no sorteio.
    amostra = amostra.sort_values(["_bairro", "_logradouro", "_id"], kind="stable").reset_index(drop=True)
    amostra["_ordem"] = amostra.index + 1

    campos = CAMPOS_ESTRUTURAL if grupo == GRUPO_ESTRUTURAL else CAMPOS_QUALIDADE
    dados = {rotulo: amostra[coluna] for rotulo, coluna in _COLUNAS_IDENTIFICACAO}
    df = pd.DataFrame(dados)
    for rotulo, _ in campos:
        df[rotulo] = ""
    return df, campos


def _aba_formulario(wb: Workbook, resultado: ResultadoAmostragem, grupo: str) -> None:
    ws = wb.active
    ws.title = "Amostra de Campo"
    df, campos = _formulario(resultado, grupo)

    titulo = (
        f"{ROTULO_GRUPO[grupo]} — {resultado.municipio or 'Município'}"
        f"{'/' + resultado.uf if resultado.uf else ''} · "
        f"{len(df)} pontos · sorteio com semente {resultado.config.semente}"
    )
    ws.cell(row=1, column=1, value=titulo)
    ws.cell(row=1, column=1).font = header_font()
    ws.cell(row=1, column=1).fill = subheader_fill()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 2))

    escrever_dataframe(ws, df, linha_inicial=2)
    aplicar_estilo_header(ws, 2, len(df.columns))
    ws.freeze_panes = "C3"
    autoajustar_largura(ws, len(df.columns))
    # A linha 1 é o título mesclado; a largura tem que sair do cabeçalho real (linha 2).
    for indice, coluna in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = max(10, min(38, len(str(coluna)) + 4))

    # Validação por lista nas colunas de campo que têm domínio fechado.
    primeira_coluna_campo = len(_COLUNAS_IDENTIFICACAO) + 1
    ultima_linha = len(df) + 2
    for deslocamento, (rotulo, opcoes) in enumerate(campos):
        if not opcoes:
            continue
        letra = get_column_letter(primeira_coluna_campo + deslocamento)
        validacao = DataValidation(
            type="list", formula1='"' + ",".join(opcoes) + '"', allow_blank=True
        )
        validacao.error = f"Valor fora da lista prevista para {rotulo}."
        validacao.errorTitle = "Valor inválido"
        ws.add_data_validation(validacao)
        validacao.add(f"{letra}3:{letra}{ultima_linha}")


def _aba_referencia(wb: Workbook, resultado: ResultadoAmostragem, grupo: str) -> None:
    """Linhas sorteadas com as colunas originais do cadastro, para conferência."""
    amostra = (resultado.estrutural if grupo == GRUPO_ESTRUTURAL else resultado.qualidade).copy()
    originais = [c for c in amostra.columns if c not in COLUNAS_AUXILIARES]
    if not originais:
        return
    ws = wb.create_sheet("Cadastro (referência)")
    df = amostra.sort_values(["_bairro", "_logradouro", "_id"], kind="stable")[originais].reset_index(drop=True)
    df.insert(0, "Nº", df.index + 1)
    escrever_dataframe(ws, df, linha_inicial=1)
    ws.freeze_panes = "B2"
    autoajustar_largura(ws, len(df.columns))


def _aba_plano(wb: Workbook, resultado: ResultadoAmostragem, grupo: str) -> None:
    """Memória de cálculo do sorteio."""
    ws = wb.create_sheet("Plano de Amostragem")
    linha = 1

    def _titulo(texto: str) -> None:
        nonlocal linha
        ws.cell(row=linha, column=1, value=texto)
        aplicar_estilo_header(ws, linha, 2)
        linha += 1

    def _item(rotulo: str, valor) -> None:
        nonlocal linha
        ws.cell(row=linha, column=1, value=rotulo)
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    _titulo("Identificação")
    _item("Município", f"{resultado.municipio}{'/' + resultado.uf if resultado.uf else ''}")
    _item("Planilha", ROTULO_GRUPO[grupo])
    _item("Pontos no cadastro (lote)", resultado.total_parque)
    _item("Pontos nesta planilha", len(resultado.estrutural if grupo == GRUPO_ESTRUTURAL else resultado.qualidade))
    _item("Amostra total (estrutural + qualidade)", resultado.total_amostra)
    _item("Semente do sorteio (reprodutibilidade)", resultado.config.semente)
    linha += 1

    plano = resultado.plano
    if plano is not None:
        _titulo("Dimensionamento — ABNT NBR 5426:1985")
        _item("Nível de inspeção", plano.nivel)
        _item("NQA (%)", plano.nqa)
        _item("Regime", plano.regime)
        _item("Letra-código (Tabela 1)", plano.letra_codigo)
        _item("Tamanho de amostra da norma", plano.tamanho_amostra)
        _item("Número de aceitação (Ac)", plano.numero_aceitacao)
        _item("Número de rejeição (Re)", plano.numero_rejeicao)
        _item("Amostra efetivamente sorteada", resultado.total_amostra)
        _item("Fração do parque inspecionada", f"{resultado.total_amostra / max(resultado.total_parque, 1):.2%}")
        for observacao in plano.observacoes:
            _item("Observação", observacao)
        linha += 1

    if not resultado.cobertura_classes.empty:
        _titulo("Cobertura por classe de iluminação")
        tabela = resultado.cobertura_classes.copy()
        tabela["% do parque"] = tabela["% do parque"].map(lambda v: f"{v:.1%}")
        linha = escrever_dataframe(ws, tabela, linha_inicial=linha) + 1

    if not resultado.cobertura_vias.empty:
        _titulo("Vias principais com cobertura obrigatória")
        linha = escrever_dataframe(ws, resultado.cobertura_vias, linha_inicial=linha) + 1

    abrangencia = resultado.abrangencia
    _titulo("Abrangência geográfica")
    _item("Bairros no cadastro / na amostra",
          f"{abrangencia.get('bairros_parque', 0)} / {abrangencia.get('bairros_amostra', 0)}")
    _item("Logradouros no cadastro / na amostra",
          f"{abrangencia.get('logradouros_parque', 0)} / {abrangencia.get('logradouros_amostra', 0)}")
    if abrangencia.get("cobertura_grid") is not None:
        _item("Células da malha 12×12 com parque atingidas pela amostra",
              f"{abrangencia['celulas_cobertas']} de {abrangencia['celulas_com_parque']} "
              f"({abrangencia['cobertura_grid']:.1%})")
        _item("Distância mediana de um ponto qualquer ao ponto inspecionado mais próximo",
              f"{abrangencia['distancia_mediana_km']:.2f} km")
        _item("Idem, percentil 90", f"{abrangencia['distancia_p90_km']:.2f} km")
    linha += 1

    if resultado.ressalvas:
        _titulo("Ressalvas")
        for ressalva in resultado.ressalvas:
            _item("•", ressalva)

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 80


def gerar(resultado: ResultadoAmostragem, grupo: str) -> bytes:
    """
    Gera a planilha .xlsx de uma das duas frentes de campo.

    Args:
        resultado: saída de `amostrador.sortear`.
        grupo: `GRUPO_ESTRUTURAL` ou `GRUPO_QUALIDADE`.

    Returns:
        Bytes do arquivo .xlsx.
    """
    if grupo not in (GRUPO_ESTRUTURAL, GRUPO_QUALIDADE):
        raise ValueError(f"Grupo desconhecido: {grupo!r}")
    wb = Workbook()
    _aba_formulario(wb, resultado, grupo)
    _aba_referencia(wb, resultado, grupo)
    _aba_plano(wb, resultado, grupo)
    return workbook_para_bytes(wb)


__all__ = ["gerar", "CAMPOS_ESTRUTURAL", "CAMPOS_QUALIDADE", "ROTULO_GRUPO"]
