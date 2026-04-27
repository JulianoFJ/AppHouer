import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('temp_simul.xlsx', data_only=True)
    if 'Simulações' in wb.sheetnames:
        sheet = wb['Simulações']
    else:
        # Try to find a sheet with 'simul' in name
        sheet_name = next((n for n in wb.sheetnames if 'simul' in n.lower()), None)
        if sheet_name:
            sheet = wb[sheet_name]
        else:
            print("Sheet not found")
            sys.exit(1)

    for r in range(1, 10):
        row_values = [str(sheet.cell(row=r, column=c).value) for c in range(1, 120)]
        print(f"Row {r}: {' | '.join(row_values[:20])} ...")
except Exception as e:
    print(f"Error: {e}")
